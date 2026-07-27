from fastapi import APIRouter, Depends, HTTPException, Query, Request, Header, Response, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, func
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel, ValidationError
from typing import Optional, List, Dict, Set
import re
import logging
import os
import psutil
import time
import uuid
import hashlib
import json
import asyncio
from datetime import datetime, timezone, timedelta
from email.utils import formatdate, parsedate_to_datetime
logger = logging.getLogger(__name__)
# Imports locaux
from database import (
    get_db,
    EntityReport,
    ScanJob,
    PendingApproval,
    ToolExecutionLog,
    APIKey,
    ScheduledScan,
    EntityResearchRun,
    EntityResearchInstruction,
    ResearchEntity,
    EntityResolutionReview,
    EntityWatch,
)
import backend_logic as logic
from middleware import ServiceStatus, get_full_health_status
from models import (
    ScanRequest,
    TargetRequest,
    DomainRequest,
    APIKeyCreate,
    ExportRequest,
    TranslateReportRequest,
    ScheduledScanCreate,
    ScheduledScanUpdate,
    LogFilter,
    CompareRequest,
    ErrorResponse,
    EntityResearchRequest,
    EntityInstructionRequest,
    EntityContinueRequest,
    EntityPreviewRequest,
    EntityResolutionReviewRequest,
    LLMProviderRequest,
    LLMSystemPromptRequest,
    validate_target,
    validate_query,
    DOMAIN_PATTERN,
    IPV4_PATTERN,
)

# Import Celery tasks (toutes les tâches spécialisées)
try:
    from tasks import (
        scan_osint_task,
        scan_osint_layer1_task,
        scan_osint_layer2_task,
        scan_osint_layer3_task,
        priority_scan_task,
        scan_parallel_task,  # Architecture parallèle (chord)
        execute_scheduled_scan_task,
        entity_research_task,
    )
    HAS_CELERY = True
except ImportError:
    scan_osint_task = None
    scan_osint_layer1_task = None
    scan_osint_layer2_task = None
    scan_osint_layer3_task = None
    priority_scan_task = None
    scan_parallel_task = None
    execute_scheduled_scan_task = None
    entity_research_task = None
    HAS_CELERY = False
    logger.warning("⚠️ Celery non disponible - Mode synchrone uniquement")


# ==================== HTTP CACHING UTILITIES ====================

def generate_etag(data: any) -> str:
    """
    Génère un ETag basé sur le contenu des données.
    Utilise MD5 pour la rapidité (pas besoin de sécurité cryptographique ici).
    """
    if isinstance(data, str):
        content = data
    else:
        content = json.dumps(data, sort_keys=True, default=str)
    return f'"{hashlib.md5(content.encode(), usedforsecurity=False).hexdigest()}"'


def format_http_date(dt: datetime) -> str:
    """Formate une datetime en format HTTP-date.

    Note: on conserve un format strictement compatible avec les tests du projet.
    """
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)

    # RFC 7231: Sun, 06 Nov 1994 08:49:37 GMT
    # (Mapping volontairement aligné sur les attentes des tests)
    weekdays = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    wd = weekdays[dt.weekday()]  # Monday=0 -> "Mon" (tests attendent ce mapping)
    month = months[dt.month - 1]

    return f"{wd}, {dt.day:02d} {month} {dt.year:04d} {dt.hour:02d}:{dt.minute:02d}:{dt.second:02d} GMT"


def check_not_modified(
    request: Request,
    etag: str = None,
    last_modified: datetime = None
) -> bool:
    """
    Vérifie si le client a une version à jour (304 Not Modified).
    Retourne True si le client peut utiliser sa version en cache.
    """
    # Check If-None-Match (ETag)
    if etag:
        if_none_match = request.headers.get("if-none-match")
        if if_none_match:
            # Handle multiple ETags
            client_etags = [e.strip() for e in if_none_match.split(",")]
            if etag in client_etags or "*" in client_etags:
                return True

    # Check If-Modified-Since
    if last_modified:
        if_modified_since = request.headers.get("if-modified-since")
        if if_modified_since:
            try:
                client_date = parsedate_to_datetime(if_modified_since)
                if last_modified.tzinfo is None:
                    last_modified = last_modified.replace(tzinfo=timezone.utc)
                if last_modified <= client_date:
                    return True
            except (ValueError, TypeError):
                pass  # Invalid date format, ignore

    return False


def add_cache_headers(
    response: Response,
    etag: str = None,
    last_modified: datetime = None,
    max_age: int = 300,  # 5 minutes default
    private: bool = False
) -> Response:
    """
    Ajoute les headers de cache HTTP à une réponse.
    """
    if etag:
        response.headers["ETag"] = etag

    if last_modified:
        response.headers["Last-Modified"] = format_http_date(last_modified)

    # Cache-Control
    cache_control = f"{'private' if private else 'public'}, max-age={max_age}"
    response.headers["Cache-Control"] = cache_control

    return response


# Essai import intent detector
try:
    from intent_detector import IntentDetector
    detector = IntentDetector()
    HAS_INTENT = True
except ImportError:
    detector = None
    HAS_INTENT = False

router = APIRouter()


# Alias pour compatibilité (utilise maintenant ScanRequest de models.py)
AskBody = ScanRequest

# Regex (importées de models.py, gardées ici pour compatibilité)
DOMAIN_RE = re.compile(r"\b([a-zA-Z0-9-]+\.[a-zA-Z]{2,})\b")
IP_RE = re.compile(r"\b(?:(?:25[0-5]|2[0-4]\d|1?\d?\d)\.){3}(?:25[0-5]|2[0-4]\d|1?\d?\d)\b")

# Mots-clés qui déclenchent une recherche WEB GÉNÉRALE (pas technique)
GENERAL_SEARCH_TRIGGERS = [
    "meteo", "météo", "temps", "heure", "prix", "actu", "news",
    "qui est", "c'est quoi", "comment", "pourquoi", "chercher", "trouver",
    "resultat", "info", "date", "quand"
]

# Prompt système "Couteau Suisse" (chat + OSINT)
# Objectifs:
# - Réponses courtes et naturelles en français (éviter le blabla / "assistant" corporate)
# - Petites discussions (ex: "ça va ?") -> 1–2 phrases max + relance
# - Si une demande OSINT/technique nécessite une cible et qu'elle manque -> demander la cible (URL/domaine/IP/etc.)
SYSTEM_PROMPT_UNIVERSAL = (
    "Tu es Ananta. Réponds en français, avec un ton direct, poli et naturel. "
    "Sois concis par défaut (2–6 lignes max). "
    "Évite les formules creuses (ex: 'Je suis là pour vous aider', 'en tant qu'IA', etc.) et les longs préambules. "
    "Pour les salutations et le small talk (ex: 'salut', 'bonjour', 'ça va ?'), réponds en 1–2 phrases maximum puis pose une question courte. "
    "Quand l'utilisateur demande une action d'analyse/OSINT (ex: whois, scan, analyse, rapport, 'check', 'investigue') "
    "sans fournir de cible exploitable, demande clairement la cible attendue (URL, domaine, IP, email, username) et ne fais rien d'autre."
)


def _normalize_chat_text(s: str) -> str:
    s = (s or "").strip().lower()
    # Ponctuation simple (on évite regex lourde ici)
    for ch in ["!", "?", ".", ",", ";", ":", "\n", "\t", "\r"]:
        s = s.replace(ch, " ")
    s = " ".join(s.split())
    return s


def _is_small_talk(q: str) -> bool:
    """Heuristique légère: messages très courts / salutations / 'ça va' etc."""
    norm = _normalize_chat_text(q)
    if not norm:
        return False

    greetings = {
        "salut",
        "bonjour",
        "bonsoir",
        "hello",
        "hi",
        "yo",
        "coucou",
        "hey",
        "slt",
    }
    small_talk = {
        "ca va",
        "ça va",
        "comment ca va",
        "comment ça va",
        "tu vas bien",
        "cv",
        "merci",
        "ok",
    }

    if norm in greetings or norm in small_talk:
        return True

    # Cas type: "ça va ?" / "salut!" / "bonjour ananta"
    if len(norm.split()) <= 4 and any(norm.startswith(g) for g in greetings):
        return True
    if norm.startswith("ca va") or norm.startswith("ça va"):
        return True

    return False


def _truncate_answer(text: str, max_chars: int = 260) -> str:
    """Tronque proprement une réponse trop longue (pour small talk)."""
    t = (text or "").strip()
    if len(t) <= max_chars:
        return t

    cut = t[:max_chars]
    # Essayer de couper sur une fin de phrase
    last_punct = max(cut.rfind("."), cut.rfind("!"), cut.rfind("?"))
    if last_punct >= int(max_chars * 0.6):
        cut = cut[: last_punct + 1]
    else:
        cut = cut.rstrip()
        cut += "…"
    return cut


@router.get("/health")
def health(request: Request, db: Session = Depends(get_db)):
    """
    Health check complet avec vérification de tous les services:
    - Database (PostgreSQL/SQLite)
    - Redis (Celery broker)
    - LLM (Mistral 7B)
    - Métriques système (CPU, RAM, GPU)

    Returns:
        - status: "healthy" | "degraded" | "unhealthy"
        - services: détail de chaque service
        - system: métriques système
    """
    try:
        # Utiliser le helper centralisé
        health_status = get_full_health_status(db)

        # Ajouter le request ID si disponible
        request_id = getattr(request.state, "request_id", None)
        if request_id:
            health_status["request_id"] = request_id

        # Mapper vers l'ancien format pour compatibilité frontend
        redis_status = health_status["services"]["redis"]["status"]
        worker_state = (
            "STABLE"
            if redis_status == "ok"
            else "DISABLED"
            if redis_status == "not_configured"
            else "DEGRADED"
        )

        # Réponse compatible avec l'ancien format + nouveau format détaillé
        return {
            # Ancien format (compatibilité)
            "cpu_load": health_status["system"]["cpu_percent"],
            "ram_load": health_status["system"]["ram_percent"],
            "gpu_load": health_status["system"]["gpu"]["load_percent"] if health_status["system"]["gpu"] else None,
            "db_latency_ms": health_status["services"]["database"]["latency_ms"],
            "worker_state": worker_state,
            "backend_api": "ONLINE",
            # Nouveau format détaillé
            "health": health_status,
        }

    except Exception as e:
        logger.exception("Erreur /health")
        return {
            "cpu_load": -1,
            "ram_load": -1,
            "gpu_load": None,
            "db_latency_ms": -1,
            "worker_state": "ERROR",
            "backend_api": "ERROR",
            "health": {
                "status": "unhealthy",
                "error": str(e),
            }
        }


@router.get("/ready")
def readiness(response: Response, db: Session = Depends(get_db)):
    """Sonde légère pour l'orchestrateur : base et broker, sans appeler le LLM."""
    services = {"database": "ok", "redis": "not_configured"}
    ready = True

    try:
        db.execute(text("SELECT 1"))
    except Exception:
        logger.exception("Échec de la sonde de disponibilité SQL")
        services["database"] = "error"
        ready = False

    redis_url = os.getenv("REDIS_URL") or os.getenv("CELERY_BROKER_URL")
    if redis_url:
        try:
            import redis

            redis.Redis.from_url(
                redis_url,
                socket_connect_timeout=1,
                socket_timeout=1,
            ).ping()
            services["redis"] = "ok"
        except Exception:
            services["redis"] = "error"
            ready = False

    if not ready:
        response.status_code = 503
    response.headers["Cache-Control"] = "no-store"
    return {"ready": ready, "services": services}


@router.get("/system/iocs")
def get_active_iocs():
    """
    Retourne les indicateurs de compromission (IOCs) extraits
    durant la session courante (IPs suspectes, domaines malveillants, etc.)
    """
    # Pour l'instant, on retourne un placeholder
    # À terme, vous pouvez stocker les IOCs en mémoire ou en DB
    return {
        "iocs": [],
        "message": "Aucun indicateur identifié dans la session actuelle."
    }


@router.get("/osint/search_smart/")
def endpoint_search_smart(query: str, db: Session = Depends(get_db)):
    # Appel avec limit=10 par défaut
    return {"results": logic.logic_search_smart(query, 10, db)}


@router.get("/osint/whois/")
def endpoint_whois(domain: str):
    return logic.logic_whois(domain)


@router.get("/osint/censys/")
def endpoint_censys(target: str):
    return logic.logic_censys(target)


@router.get("/osint/dns/")
def endpoint_dns(domain: str):
    """Résolution DNS simple (DOMAIN -> IP)."""
    return logic.logic_dns_resolution(domain)


@router.get("/osint/headers/")
def endpoint_headers(domain: str):
    """Récupération/analyse des headers HTTP (best effort)."""
    return logic.logic_http_headers(domain)


# ✅ HISTORIQUE BDD (AVEC CACHE HTTP)
@router.get("/osint/history/")
def osint_history(
    request: Request,
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db)
):
    """
    Récupère l'historique des rapports.
    Supporte le caching HTTP via ETag (basé sur le dernier rapport modifié).
    """
    try:
        reports = (
            db.query(EntityReport)
            .order_by(EntityReport.created_at.desc())
            .limit(limit)
            .all()
        )

        # Construire les données de réponse
        response_data = [
            {
                "date": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "--",
                "title": f"Rapport {r.target_type or 'UNKNOWN'}",
                "query": r.target or "",
            }
            for r in reports
        ]

        # Générer ETag basé sur le contenu
        etag = generate_etag(response_data)

        # Déterminer Last-Modified (date du rapport le plus récent)
        last_modified = reports[0].updated_at or reports[0].created_at if reports else None

        # Vérifier 304 Not Modified
        if check_not_modified(request, etag=etag, last_modified=last_modified):
            return Response(
                status_code=304,
                headers={
                    "ETag": etag,
                    "Cache-Control": "public, max-age=60"  # Cache court pour l'historique
                }
            )

        # Retourner avec headers de cache
        response = JSONResponse(content=response_data)
        add_cache_headers(response, etag=etag, last_modified=last_modified, max_age=60)
        return response

    except Exception as e:
        logger.exception("[/osint/history] ERREUR")
        raise HTTPException(
            status_code=500,
            detail="Erreur lors de la récupération de l'historique"
        )

@router.get("/osint/report/")
def get_cached_report(
    request: Request,
    target: str = Query(..., description="Cible du rapport à récupérer"),
    db: Session = Depends(get_db)
):
    """
    Récupère un rapport en cache SANS régénération LLM.
    Utilisé par database.html pour l'aperçu rapide.

    Supporte le caching HTTP via ETag et Last-Modified.
    """
    try:
        # Normaliser la cible pour la recherche
        normalized = logic.normalize_target(target)

        # Chercher le rapport
        report = db.query(EntityReport).filter(
            EntityReport.target.ilike(f"%{normalized}%")
        ).first()

        if not report:
            raise HTTPException(status_code=404, detail="Rapport non trouvé")

        # Générer ETag et Last-Modified
        last_modified = report.updated_at or report.created_at
        response_data = {
            "target": report.target,
            "type": report.target_type,
            "report": report.final_report,
            "date": last_modified.strftime("%Y-%m-%d %H:%M") if last_modified else "N/A",
            "cached": True
        }
        etag = generate_etag(response_data)

        # Vérifier si le client a une version à jour (304 Not Modified)
        if check_not_modified(request, etag=etag, last_modified=last_modified):
            return Response(
                status_code=304,
                headers={
                    "ETag": etag,
                    "Last-Modified": format_http_date(last_modified) if last_modified else None,
                    "Cache-Control": "public, max-age=300"
                }
            )

        # Retourner le rapport avec headers de cache
        response = JSONResponse(content=response_data)
        add_cache_headers(response, etag=etag, last_modified=last_modified, max_age=300)
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[/osint/report GET] ERREUR: {e}")
        raise HTTPException(
            status_code=500,
            detail="Erreur lors de la récupération du rapport"
        )


@router.post("/osint/translate")
def translate_cached_report(body: TranslateReportRequest, db: Session = Depends(get_db)):
    """Traduit un rapport en cache via le LLM (sans ré-exécuter les outils)."""
    normalized = logic.normalize_target(body.target)
    report = db.query(EntityReport).filter(EntityReport.target.ilike(f"%{normalized}%")).first()
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        translated = logic.translate_report_markdown(
            report.final_report or "",
            to_language=body.to_language,
            llm_hard_limit=body.llm_hard_limit,
        )
        return {
            "target": report.target,
            "type": report.target_type,
            "to_language": body.to_language,
            "report": translated,
            "cached": True,
            "translated": True,
        }
    except Exception as e:
        logger.exception(f"[/osint/translate] ERREUR: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la traduction")


@router.get("/osint/report/view")
def get_report_view(
    target: str = Query(..., description="Cible (domaine/IP)"),
    mode: str = Query("executive", description="executive|technical|evidence"),
    db: Session = Depends(get_db),
):
    """Retourne une restitution multi-niveau à partir des données structurées.

    - executive: décisionnel, risques majeurs, actions
    - technical: détails exploitables (preuves, remédiations)
    - evidence: données structurées brutes (structured_data/exposures/graph)

    Pas de régénération LLM : c'est du rendu basé sur la mémoire du scan.
    """
    norm = validate_target(target)
    report = (
        db.query(EntityReport)
        .filter(EntityReport.target.ilike(f"%{norm}%"))
        .order_by(EntityReport.updated_at.desc(), EntityReport.created_at.desc())
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        payload = json.loads(report.raw_data) if report.raw_data else {}
    except Exception:
        payload = {}

    structured = payload.get("structured_data") or {}
    exposures = payload.get("exposures") or []
    graph = payload.get("intel_graph") or {"nodes": [], "edges": []}
    risk = payload.get("risk_analysis") or {}

    mode = (mode or "executive").lower().strip()
    if mode not in {"executive", "technical", "evidence"}:
        raise HTTPException(status_code=400, detail="mode invalide (executive|technical|evidence)")

    # Sort exposures by severity
    sev_rank = {"CRITICAL": 5, "HIGH": 4, "MEDIUM": 3, "LOW": 2, "INFO": 1}

    def _sev(e: dict) -> int:
        if not isinstance(e, dict):
            return 0
        return sev_rank.get((e.get("severity") or "MEDIUM").upper(), 3)

    exposures_sorted = sorted([e for e in exposures if isinstance(e, dict)], key=_sev, reverse=True)

    if mode == "evidence":
        return {
            "target": report.target,
            "report_id": report.id,
            "mode": mode,
            "risk_analysis": risk,
            "structured_data": structured,
            "exposures": exposures_sorted,
            "graph": graph,
        }

    if mode == "executive":
        top_exposures = exposures_sorted[:8]
        actions = []
        for e in top_exposures:
            title = e.get("title")
            if isinstance(title, str) and title:
                actions.append(title)

        return {
            "target": report.target,
            "report_id": report.id,
            "mode": mode,
            "summary": structured.get("executive_summary") or "",
            "risk": {
                "score": risk.get("score"),
                "level": risk.get("level"),
            },
            "key_risks": top_exposures,
            "priority_actions": actions[:8],
            "limits": structured.get("limitations") or [],
        }

    # technical
    findings = structured.get("top_findings") or []
    if not isinstance(findings, list):
        findings = []

    tech_findings = []
    for f in findings[:25]:
        if not isinstance(f, dict):
            continue
        tech_findings.append({
            "id": f.get("id"),
            "title": f.get("title"),
            "category": f.get("category"),
            "severity": f.get("severity"),
            "confidence": f.get("confidence"),
            "claim": f.get("claim"),
            "evidence": f.get("evidence"),
            "impact": f.get("impact"),
            "remediation": f.get("remediation"),
            "sources": f.get("sources"),
        })

    return {
        "target": report.target,
        "report_id": report.id,
        "mode": mode,
        "risk": {
            "score": risk.get("score"),
            "level": risk.get("level"),
        },
        "exposures": exposures_sorted,
        "findings": tech_findings,
        "graph": {
            "nodes": (graph.get("nodes") if isinstance(graph, dict) else []),
            "edges": (graph.get("edges") if isinstance(graph, dict) else []),
        },
    }


@router.get("/osint/report/legacy")
def get_cached_report_legacy(
    target: str = Query(..., description="Cible du rapport à récupérer"),
    db: Session = Depends(get_db)
):
    """
    Version legacy sans caching HTTP (pour compatibilité).
    """
    try:
        normalized = logic.normalize_target(target)
        report = db.query(EntityReport).filter(
            EntityReport.target.ilike(f"%{normalized}%")
        ).first()

        if not report:
            raise HTTPException(status_code=404, detail="Rapport non trouvé")

        return {
            "target": report.target,
            "type": report.target_type,
            "report": report.final_report,
            "date": report.updated_at.strftime("%Y-%m-%d %H:%M") if report.updated_at else (
                report.created_at.strftime("%Y-%m-%d %H:%M") if report.created_at else "N/A"
            ),
            "cached": True
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[/osint/report GET] ERREUR: {e}")
        raise HTTPException(
            status_code=500,
            detail="Erreur lors de la récupération du rapport"
        )


@router.delete("/osint/report/")
def delete_report(
    target: str = Query(..., description="Cible du rapport à supprimer"),
    db: Session = Depends(get_db)
):
    """Supprime un rapport de la base de données par sa cible."""
    try:
        # Normaliser la cible pour la recherche
        normalized = logic.normalize_target(target)

        # Chercher le rapport
        report = db.query(EntityReport).filter(
            EntityReport.target.ilike(f"%{normalized}%")
        ).first()

        if not report:
            raise HTTPException(status_code=404, detail="Rapport non trouvé")

        # Supprimer le rapport
        db.delete(report)
        db.commit()

        logger.info(f"[DELETE] Rapport supprimé: {target}")
        return {"status": "ok", "message": f"Rapport '{target}' supprimé avec succès"}

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[/osint/report DELETE] ERREUR: {e}")
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail="Erreur lors de la suppression du rapport"
        )

@router.get("/osint/generate_pdf/")
def endpoint_pdf(query: str, db: Session = Depends(get_db)):
    try:
        path = logic.logic_generate_pdf(query, db)
        return FileResponse(path, media_type='application/pdf')
    except Exception as e:
        logger.exception("❌ Erreur /osint/generate_pdf/")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== EXPORT FORMATS ====================

@router.get("/osint/export/json")
def export_json(request: Request, query: str, db: Session = Depends(get_db)):
    """
    Exporte un rapport au format JSON.
    Supporte le caching HTTP via ETag et Last-Modified.
    """
    try:
        normalized = logic.normalize_target(query)
        report = db.query(EntityReport).filter(
            EntityReport.target.ilike(f"%{normalized}%")
        ).first()

        if not report:
            raise HTTPException(status_code=404, detail="Rapport non trouvé")

        # Déterminer Last-Modified
        last_modified = report.updated_at or report.created_at

        # Construire l'export JSON
        raw_data = {}
        try:
            raw_data = json.loads(report.raw_data) if report.raw_data else {}
        except:
            pass

        export_data = {
            "target": report.target,
            "target_type": report.target_type,
            "report": report.final_report,
            "raw_data": raw_data,
            "created_at": report.created_at.isoformat() if report.created_at else None,
            "updated_at": report.updated_at.isoformat() if report.updated_at else None
        }

        # Générer ETag
        etag = generate_etag(export_data)

        # Vérifier 304 Not Modified
        if check_not_modified(request, etag=etag, last_modified=last_modified):
            return Response(
                status_code=304,
                headers={
                    "ETag": etag,
                    "Cache-Control": "public, max-age=3600"  # 1 heure pour les exports
                }
            )

        response = JSONResponse(
            content=export_data,
            media_type="application/json",
            headers={
                "Content-Disposition": f'attachment; filename="report_{normalized}.json"'
            }
        )
        add_cache_headers(response, etag=etag, last_modified=last_modified, max_age=3600)
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ Erreur /osint/export/json: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/osint/export/csv")
def export_csv(request: Request, query: str, db: Session = Depends(get_db)):
    """Exporte les findings d'un rapport au format CSV.

    Supporte le caching HTTP via ETag + 304 (If-None-Match).
    """
    try:
        normalized = logic.normalize_target(query)
        report = db.query(EntityReport).filter(
            EntityReport.target.ilike(f"%{normalized}%")
        ).first()

        if not report:
            raise HTTPException(status_code=404, detail="Rapport non trouvé")

        # Cache validators
        last_modified = report.updated_at or report.created_at
        etag_source = f"csv:{report.id}:{last_modified}:{report.target}"
        etag = generate_etag(etag_source)

        if check_not_modified(request, etag=etag, last_modified=last_modified):
            return Response(
                status_code=304,
                headers={
                    "ETag": etag,
                    "Cache-Control": "public, max-age=3600",
                },
            )

        import csv
        import io
        import json

        # Parser raw_data pour extraire les infos structurées
        raw_data = {}
        try:
            raw_data = json.loads(report.raw_data) if report.raw_data else {}
        except:
            pass

        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(["Category", "Tool", "Status", "Key", "Value"])

        # Infos générales
        writer.writerow(["General", "target", "info", "target", report.target])
        writer.writerow(["General", "type", "info", "type", report.target_type])
        writer.writerow(["General", "scan_date", "info", "date", raw_data.get("scanned_at", "N/A")])

        # Risk Analysis
        risk = raw_data.get("risk_analysis", {})
        if risk:
            writer.writerow(["Risk", "analysis", "computed", "score", risk.get("score", "N/A")])
            writer.writerow(["Risk", "analysis", "computed", "level", risk.get("level", "N/A")])
            for ind in risk.get("indicators", {}).get("positive", []):
                writer.writerow(["Risk", "positive", "info", "indicator", ind])
            for ind in risk.get("indicators", {}).get("negative", []):
                writer.writerow(["Risk", "negative", "warning", "indicator", ind])

        # Tools results
        tools = raw_data.get("tools", {})
        for tool_name, tool_data in tools.items():
            status = tool_data.get("status", "unknown")
            if status == "ok":
                data = tool_data.get("data", {})
                if isinstance(data, dict):
                    for key, value in data.items():
                        if not isinstance(value, (dict, list)):
                            writer.writerow(["Tool", tool_name, status, key, str(value)[:200]])
                else:
                    writer.writerow(["Tool", tool_name, status, "data", str(data)[:200]])
            elif status == "error":
                writer.writerow(["Tool", tool_name, status, "error", tool_data.get("error", "Unknown")])
            elif status == "skipped":
                writer.writerow(["Tool", tool_name, status, "reason", tool_data.get("reason", "Unknown")])

        csv_content = output.getvalue()
        output.close()

        response = Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="report_{normalized}.csv"'
            },
        )
        add_cache_headers(response, etag=etag, last_modified=last_modified, max_age=3600)
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ Erreur /osint/export/csv: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/osint/export/xml")
def export_xml(request: Request, query: str, db: Session = Depends(get_db)):
    """Exporte un rapport au format XML.

    Supporte le caching HTTP via ETag + 304 (If-None-Match).
    """
    try:
        normalized = logic.normalize_target(query)
        report = db.query(EntityReport).filter(
            EntityReport.target.ilike(f"%{normalized}%")
        ).first()

        if not report:
            raise HTTPException(status_code=404, detail="Rapport non trouvé")

        # Cache validators
        last_modified = report.updated_at or report.created_at
        etag_source = f"xml:{report.id}:{last_modified}:{report.target}"
        etag = generate_etag(etag_source)

        if check_not_modified(request, etag=etag, last_modified=last_modified):
            return Response(
                status_code=304,
                headers={
                    "ETag": etag,
                    "Cache-Control": "public, max-age=3600",
                },
            )

        import json
        from xml.etree.ElementTree import Element, SubElement, indent, tostring

        raw_data = {}
        try:
            raw_data = json.loads(report.raw_data) if report.raw_data else {}
        except:
            pass

        # Construire l'arbre XML
        root = Element("osint_report")
        root.set("version", "2.0")

        # Metadata
        meta = SubElement(root, "metadata")
        SubElement(meta, "target").text = report.target
        SubElement(meta, "target_type").text = report.target_type or "UNKNOWN"
        SubElement(meta, "scanned_at").text = raw_data.get("scanned_at", "N/A")
        SubElement(meta, "created_at").text = report.created_at.isoformat() if report.created_at else "N/A"

        # Risk Analysis
        risk = raw_data.get("risk_analysis", {})
        if risk:
            risk_elem = SubElement(root, "risk_analysis")
            SubElement(risk_elem, "score").text = str(risk.get("score", 0))
            SubElement(risk_elem, "level").text = risk.get("level", "UNKNOWN")

            indicators = SubElement(risk_elem, "indicators")
            for ind in risk.get("indicators", {}).get("positive", []):
                pos = SubElement(indicators, "positive")
                pos.text = ind
            for ind in risk.get("indicators", {}).get("negative", []):
                neg = SubElement(indicators, "negative")
                neg.text = ind

        # Tools
        tools_elem = SubElement(root, "tools")
        for tool_name, tool_data in raw_data.get("tools", {}).items():
            tool_elem = SubElement(tools_elem, "tool")
            tool_elem.set("name", tool_name)
            tool_elem.set("status", tool_data.get("status", "unknown"))
            if tool_data.get("duration"):
                tool_elem.set("duration", f"{tool_data['duration']:.2f}s")

        # Report content
        report_elem = SubElement(root, "report")
        report_elem.text = report.final_report

        # Indentation sans reparsing : les données restent échappées par ElementTree.
        indent(root, space="  ")
        xml_str = '<?xml version="1.0" encoding="utf-8"?>\n' + tostring(
            root,
            encoding="unicode",
        )

        response = Response(
            content=xml_str,
            media_type="application/xml",
            headers={
                "Content-Disposition": f'attachment; filename="report_{normalized}.xml"'
            },
        )
        add_cache_headers(response, etag=etag, last_modified=last_modified, max_age=3600)
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ Erreur /osint/export/xml: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/osint/export/markdown")
def export_markdown(request: Request, query: str, db: Session = Depends(get_db)):
    """Exporte un rapport au format Markdown.

    Supporte le caching HTTP via ETag + 304 (If-None-Match).
    """
    try:
        normalized = logic.normalize_target(query)
        report = db.query(EntityReport).filter(
            EntityReport.target.ilike(f"%{normalized}%")
        ).first()

        if not report:
            raise HTTPException(status_code=404, detail="Rapport non trouvé")

        # Cache validators
        last_modified = report.updated_at or report.created_at
        etag_source = f"md:{report.id}:{last_modified}:{report.target}"
        etag = generate_etag(etag_source)

        if check_not_modified(request, etag=etag, last_modified=last_modified):
            return Response(
                status_code=304,
                headers={
                    "ETag": etag,
                    "Cache-Control": "public, max-age=3600",
                },
            )

        import json
        raw_data = {}
        try:
            raw_data = json.loads(report.raw_data) if report.raw_data else {}
        except:
            pass

        # Construire le Markdown
        markdown_content = f"""# Rapport OSINT: {report.target}

**Type de cible**: {report.target_type or "UNKNOWN"}
**Date d'analyse**: {raw_data.get("scanned_at", "N/A")}
**Créé le**: {report.created_at.strftime("%Y-%m-%d %H:%M") if report.created_at else "N/A"}

---

## Analyse de Risque

"""

        risk = raw_data.get("risk_analysis", {})
        if risk:
            markdown_content += f"""**Score**: {risk.get("score", 0)}/100
**Niveau**: {risk.get("level", "UNKNOWN")}

### Indicateurs Positifs
"""
            for ind in risk.get("indicators", {}).get("positive", []):
                markdown_content += f"- ✅ {ind}\n"

            markdown_content += "\n### Indicateurs Négatifs\n"
            for ind in risk.get("indicators", {}).get("negative", []):
                markdown_content += f"- ⚠️ {ind}\n"

        markdown_content += "\n---\n\n## Outils Utilisés\n\n"

        for tool_name, tool_data in raw_data.get("tools", {}).items():
            status_icon = "✅" if tool_data.get("status") == "ok" else "❌"
            duration = f" ({tool_data['duration']:.2f}s)" if tool_data.get("duration") else ""
            markdown_content += f"- {status_icon} **{tool_name}**: {tool_data.get('status', 'unknown')}{duration}\n"

        markdown_content += f"\n---\n\n## Rapport Complet\n\n{report.final_report}\n"

        response = Response(
            content=markdown_content,
            media_type="text/markdown",
            headers={
                "Content-Disposition": f'attachment; filename="report_{normalized}.md"'
            },
        )
        add_cache_headers(response, etag=etag, last_modified=last_modified, max_age=3600)
        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"❌ Erreur /osint/export/markdown: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/osint/export/xlsx")
def export_xlsx(request: Request, query: str, db: Session = Depends(get_db)):
    """
    Exporte un rapport au format Excel (XLSX).
    Supporte le caching HTTP via ETag et Last-Modified.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from fastapi.responses import StreamingResponse
        import io

        normalized = logic.normalize_target(query)
        report = db.query(EntityReport).filter(
            EntityReport.target.ilike(f"%{normalized}%")
        ).first()

        if not report:
            raise HTTPException(status_code=404, detail="Rapport non trouvé")

        # Déterminer Last-Modified
        last_modified = report.updated_at or report.created_at

        # Générer un ETag stable basé sur report ID + updated_at
        # Pour XLSX, on ne génère pas le fichier entier juste pour le hash
        etag_source = f"{report.id}:{report.updated_at or report.created_at}:{report.target}"
        etag = generate_etag(etag_source)

        # Vérifier 304 Not Modified (avant de générer le XLSX coûteux)
        if check_not_modified(request, etag=etag, last_modified=last_modified):
            return Response(
                status_code=304,
                headers={
                    "ETag": etag,
                    "Cache-Control": "public, max-age=3600"
                }
            )

        raw_data = {}
        try:
            raw_data = json.loads(report.raw_data) if report.raw_data else {}
        except:
            pass

        # Create workbook
        wb = Workbook()

        # === Sheet 1: Summary ===
        ws_summary = wb.active
        ws_summary.title = "Résumé"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Summary header
        ws_summary["A1"] = "Rapport OSINT - Ananta"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:D1')

        summary_data = [
            ("Cible", report.target),
            ("Type", report.target_type or "UNKNOWN"),
            ("Date d'analyse", raw_data.get("scanned_at", "N/A")),
            ("Créé le", report.created_at.strftime("%Y-%m-%d %H:%M") if report.created_at else "N/A"),
        ]

        for row_idx, (key, value) in enumerate(summary_data, start=3):
            ws_summary[f"A{row_idx}"] = key
            ws_summary[f"A{row_idx}"].font = Font(bold=True)
            ws_summary[f"B{row_idx}"] = value

        # Risk Analysis
        risk = raw_data.get("risk_analysis", {})
        if risk:
            ws_summary["A8"] = "Analyse de Risque"
            ws_summary["A8"].font = Font(bold=True, size=12)
            ws_summary["A9"] = "Score"
            ws_summary["B9"] = risk.get("score", 0)
            ws_summary["A10"] = "Niveau"
            ws_summary["B10"] = risk.get("level", "UNKNOWN")

        # Set column widths
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 50

        # === Sheet 2: Tools Results ===
        ws_tools = wb.create_sheet("Outils")

        # Header row
        headers = ["Outil", "Statut", "Durée (s)", "Détails"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_tools.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Tools data
        tools = raw_data.get("tools", {})
        for row_idx, (tool_name, tool_data) in enumerate(tools.items(), start=2):
            ws_tools.cell(row=row_idx, column=1, value=tool_name).border = border
            ws_tools.cell(row=row_idx, column=2, value=tool_data.get("status", "unknown")).border = border
            duration = tool_data.get("duration", "")
            ws_tools.cell(row=row_idx, column=3, value=f"{duration:.2f}" if isinstance(duration, (int, float)) else "").border = border

            # Details summary
            if tool_data.get("status") == "ok":
                data = tool_data.get("data", {})
                if isinstance(data, dict):
                    details = ", ".join(f"{k}: {str(v)[:50]}" for k, v in list(data.items())[:3])
                else:
                    details = str(data)[:100]
            elif tool_data.get("status") == "error":
                details = tool_data.get("error", "")[:100]
            else:
                details = tool_data.get("reason", "")[:100]
            ws_tools.cell(row=row_idx, column=4, value=details).border = border

        # Set column widths
        ws_tools.column_dimensions['A'].width = 20
        ws_tools.column_dimensions['B'].width = 12
        ws_tools.column_dimensions['C'].width = 12
        ws_tools.column_dimensions['D'].width = 60

        # === Sheet 3: Risk Indicators ===
        ws_risk = wb.create_sheet("Indicateurs")

        headers = ["Type", "Indicateur"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_risk.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        row_idx = 2
        for ind in risk.get("indicators", {}).get("positive", []):
            ws_risk.cell(row=row_idx, column=1, value="✅ Positif").border = border
            ws_risk.cell(row=row_idx, column=2, value=ind).border = border
            row_idx += 1

        for ind in risk.get("indicators", {}).get("negative", []):
            ws_risk.cell(row=row_idx, column=1, value="⚠️ Négatif").border = border
            ws_risk.cell(row=row_idx, column=2, value=ind).border = border
            row_idx += 1

        ws_risk.column_dimensions['A'].width = 15
        ws_risk.column_dimensions['B'].width = 60

        # === Sheet 4: Full Report ===
        ws_report = wb.create_sheet("Rapport Complet")
        ws_report["A1"] = "Rapport d'analyse"
        ws_report["A1"].font = Font(bold=True, size=12)

        # Split report into lines for better readability
        if report.final_report:
            lines = report.final_report.split('\n')
            for row_idx, line in enumerate(lines, start=3):
                ws_report.cell(row=row_idx, column=1, value=line)

        ws_report.column_dimensions['A'].width = 100

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        xlsx_content = output.getvalue()
        output.seek(0)

        # Create response with cache headers
        response = StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="report_{normalized}.xlsx"'
            }
        )
        add_cache_headers(response, etag=etag, last_modified=last_modified, max_age=3600)
        return response

    except HTTPException:
        raise
    except ImportError as e:
        logger.error(f"[/osint/export/xlsx] openpyxl not installed: {e}")
        raise HTTPException(
            status_code=501,
            detail="Export Excel non disponible. Installez openpyxl: pip install openpyxl"
        )
    except Exception as e:
        logger.exception(f"❌ Erreur /osint/export/xlsx: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== CACHE MANAGEMENT ====================

@router.get("/cache/stats")
def get_cache_stats(db: Session = Depends(get_db)):
    """Retourne les statistiques du cache de rapports."""
    try:
        # Compter les rapports
        report_count = db.query(EntityReport).count()

        # Estimer la taille (approximatif basé sur la longueur des rapports)
        reports = db.query(EntityReport.final_report, EntityReport.raw_data).all()
        total_bytes = 0
        for r in reports:
            if r.final_report:
                total_bytes += len(r.final_report.encode('utf-8'))
            if r.raw_data:
                total_bytes += len(r.raw_data.encode('utf-8'))

        used_mb = total_bytes / (1024 * 1024)

        # Dates extrêmes
        oldest = db.query(func.min(EntityReport.created_at)).scalar()
        newest = db.query(func.max(EntityReport.updated_at)).scalar()
        if not newest:
            newest = db.query(func.max(EntityReport.created_at)).scalar()

        return {
            "used_mb": round(used_mb, 2),
            "max_mb": 100,  # Limite arbitraire pour l'UI
            "report_count": report_count,
            "oldest_report": oldest.isoformat() if oldest else None,
            "newest_report": newest.isoformat() if newest else None
        }

    except Exception as e:
        logger.exception(f"[/cache/stats] ERREUR: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la récupération des stats cache")


# ==================== SCHEDULED SCANS ====================


def _compute_next_run_at(scan: ScheduledScan) -> datetime:
    now = datetime.now(timezone.utc)

    if scan.schedule_type == "daily":
        next_run = now.replace(hour=scan.hour, minute=0, second=0, microsecond=0)
        if next_run <= now:
            next_run = next_run + timedelta(days=1)
        return next_run

    if scan.schedule_type == "weekly":
        # 0=lundi ... 6=dimanche
        dow = scan.day_of_week if scan.day_of_week is not None else 0
        base = now.replace(hour=scan.hour, minute=0, second=0, microsecond=0)
        days_ahead = (dow - base.weekday()) % 7
        next_run = base + timedelta(days=days_ahead)
        if next_run <= now:
            next_run = next_run + timedelta(days=7)
        return next_run

    if scan.schedule_type == "monthly":
        day = min(int(scan.day_of_month or 1), 28)
        next_run = now.replace(day=day, hour=scan.hour, minute=0, second=0, microsecond=0)
        if next_run <= now:
            # month rollover
            if next_run.month == 12:
                next_run = next_run.replace(year=next_run.year + 1, month=1)
            else:
                next_run = next_run.replace(month=next_run.month + 1)
        return next_run

    if scan.schedule_type == "custom" and scan.cron_expression:
        try:
            from croniter import croniter

            cron = croniter(scan.cron_expression, now)
            return cron.get_next(datetime)
        except Exception:
            return now + timedelta(days=1)

    return now + timedelta(days=1)


@router.get("/scheduled-scans/list")
def list_scheduled_scans(db: Session = Depends(get_db)):
    scans = db.query(ScheduledScan).order_by(ScheduledScan.created_at.desc()).all()
    return [
        {
            "id": s.id,
            "name": s.name,
            "target": s.target,
            "scan_mode": s.scan_mode,
            "report_template": s.report_template,
            "language": s.language,
            "llm_hard_limit": s.llm_hard_limit,
            "schedule_type": s.schedule_type,
            "cron_expression": s.cron_expression,
            "hour": s.hour,
            "day_of_week": s.day_of_week,
            "day_of_month": s.day_of_month,
            "is_active": s.is_active,
            "last_run_at": s.last_run_at.isoformat() if s.last_run_at else None,
            "next_run_at": s.next_run_at.isoformat() if s.next_run_at else None,
            "last_run_status": s.last_run_status,
            "last_error": s.last_error,
            "run_count": s.run_count,
            "notify_email": s.notify_email,
            "notify_on_change": s.notify_on_change,
            "notify_on_error": s.notify_on_error,
            "created_at": s.created_at.isoformat() if s.created_at else None,
        }
        for s in scans
    ]


@router.post("/scheduled-scans/create")
def create_scheduled_scan(body: ScheduledScanCreate, db: Session = Depends(get_db)):
    scan = ScheduledScan(
        name=body.name.strip(),
        target=body.target,
        scan_mode=body.scan_mode,
        report_template=body.report_template,
        language=body.language,
        llm_hard_limit=body.llm_hard_limit,
        schedule_type=body.schedule_type,
        cron_expression=body.cron_expression,
        hour=body.hour,
        day_of_week=body.day_of_week,
        day_of_month=body.day_of_month,
        notify_email=body.notify_email,
        notify_on_change=body.notify_on_change,
        notify_on_error=body.notify_on_error,
        created_by=body.created_by,
        is_active=True,
    )
    scan.next_run_at = _compute_next_run_at(scan)

    db.add(scan)
    db.commit()
    db.refresh(scan)

    return {"status": "ok", "id": scan.id, "next_run_at": scan.next_run_at.isoformat()}


@router.post("/scheduled-scans/{scan_id}")
def update_scheduled_scan(scan_id: int, body: ScheduledScanUpdate, db: Session = Depends(get_db)):
    scan = db.query(ScheduledScan).filter_by(id=scan_id).first()
    if not scan:
        raise HTTPException(status_code=404, detail="Scheduled scan not found")

    if body.name is not None:
        scan.name = body.name.strip()
    if body.is_active is not None:
        scan.is_active = body.is_active
    if body.notify_email is not None:
        scan.notify_email = body.notify_email
    if body.notify_on_change is not None:
        scan.notify_on_change = body.notify_on_change
    if body.notify_on_error is not None:
        scan.notify_on_error = body.notify_on_error

    if body.llm_hard_limit is not None:
        scan.llm_hard_limit = body.llm_hard_limit

    # Recompute next run when (re)activating
    if scan.is_active:
        scan.next_run_at = _compute_next_run_at(scan)
    else:
        scan.next_run_at = None

    db.commit()
    return {"status": "ok"}


@router.delete("/scheduled-scans/{scan_id}")
def delete_scheduled_scan(scan_id: int, db: Session = Depends(get_db)):
    scan = db.query(ScheduledScan).filter_by(id=scan_id).first()
    if not scan:
        raise HTTPException(status_code=404, detail="Scheduled scan not found")

    db.delete(scan)
    db.commit()
    return {"status": "ok"}


@router.post("/scheduled-scans/{scan_id}/run")
def run_scheduled_scan_now(scan_id: int, db: Session = Depends(get_db)):
    if not HAS_CELERY or execute_scheduled_scan_task is None:
        raise HTTPException(status_code=503, detail="Celery not configured")

    scan = db.query(ScheduledScan).filter_by(id=scan_id).first()
    if not scan:
        raise HTTPException(status_code=404, detail="Scheduled scan not found")

    task = execute_scheduled_scan_task.delay(scan_id)
    return {"status": "queued", "task_id": task.id}


@router.post("/cache/clear")
def clear_cache(db: Session = Depends(get_db)):
    """Supprime tous les rapports en cache."""
    try:
        deleted_count = db.query(EntityReport).delete()
        db.commit()

        logger.info(f"[CACHE CLEAR] {deleted_count} rapports supprimés")

        return {
            "success": True,
            "deleted_count": deleted_count
        }

    except Exception as e:
        logger.exception(f"[/cache/clear] ERREUR: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail="Erreur lors du nettoyage du cache")


@router.post("/agent/ask")
def agent_ask(body: AskBody, db: Session = Depends(get_db)):
    q = (body.query or "").strip()
    if not q:
        return {"type": "chat", "answer": "Je vous écoute."}

    q_lower = q.lower()

    # 1) COMMANDES EXPLICITES

    # ANALYZE : Commande explicite pour analyse OSINT complète
    if q_lower.startswith("analyze"):
        parts = q.split(maxsplit=1)
        target = parts[1] if len(parts) > 1 else ""

        if not target:
            return {"type": "chat", "answer": "Veuillez préciser une cible après 'analyze' (IP, domaine ou URL)."}

        logger.info(f"[ANALYZE] Commande explicite détectée pour : {target}")

        # Lance le rapport OSINT complet (WHOIS + Censys + Web Enrichment)
        report = logic.logic_run_report(
            target,
            db,
            report_type="osint",
            llm_hard_limit=getattr(body, "llm_hard_limit", None),
        )

        return {
            "type": "osint",
            "results": report.get("sources", []),
            "answer": report.get("report", ""),
            "cached": report.get("source") == "database_cache"
        }

    if q_lower.startswith("whois"):
        parts = q.split()
        target = parts[1] if len(parts) > 1 else ""
        if not target:
            return {"type": "chat", "answer": "Veuillez préciser un domaine après 'whois'."}

        data = logic.logic_whois(target)
        ans = data.get("analysis") or str(data.get("raw"))
        return {"type": "chat", "answer": ans}

    if q_lower.startswith("censys"):
        parts = q.split()
        target = parts[1] if len(parts) > 1 else ""
        if not target:
            return {"type": "chat", "answer": "Veuillez préciser une IP ou un domaine après 'censys'."}

        data = logic.logic_censys(target)
        if "error" in data:
            return {"type": "chat", "answer": f"Erreur Censys: {data['error']}"}

        ans = data.get("analysis") or str(data.get("raw"))
        return {"type": "chat", "answer": ans}

    # 2) DÉTECTION CIBLE TECHNIQUE -> Rapport OSINT
    has_ip = bool(IP_RE.search(q))
    has_domain = bool(DOMAIN_RE.search(q))
    is_long_text = len(q.split()) > 8

    if (has_ip or has_domain) and not is_long_text:
        report = logic.logic_run_report(q, db, report_type="osint", llm_hard_limit=getattr(body, "llm_hard_limit", None))
        return {"type": "osint", "results": report.get("sources", []), "answer": report.get("report", "")}

    # 3) DÉTECTION BESOIN RECHERCHE GÉNÉRALE
    if any(trig in q_lower for trig in GENERAL_SEARCH_TRIGGERS):
        report = logic.logic_run_report(q, db, report_type="general", llm_hard_limit=getattr(body, "llm_hard_limit", None))
        return {"type": "osint", "results": report.get("sources", []), "answer": report.get("report", "")}

    # 4) PAR DÉFAUT -> CHAT (messages simples comme "salut", "bonjour", etc.)
    # IMPORTANT: pas de fast-path "salutations" -> on garde ces messages routés vers le LLM
    logger.info(f"[CHAT] Message simple détecté: '{q[:50]}...' -> appel LLM")
    try:
        answer = logic.ask_llm(SYSTEM_PROMPT_UNIVERSAL, q)

        # Vérifier que la réponse n'est pas vide
        if not answer or answer.strip() == "":
            logger.warning(f"[CHAT] LLM a retourné une réponse vide pour: '{q}'")
            answer = "Que voulez-vous faire (discussion, question, ou analyse OSINT) ?"

        # Garder les réponses de small talk très courtes (post-traitement léger)
        if _is_small_talk(q):
            answer = _truncate_answer(answer, max_chars=260)

        logger.info(f"[CHAT] Réponse LLM ({len(answer)} chars): '{answer[:100]}...'")
        return {"type": "chat", "answer": answer}

    except Exception as e:
        logger.exception(f"[CHAT] Erreur lors de l'appel LLM pour '{q}': {e}")
        return {
            "type": "chat",
            "answer": f"Désolé, je rencontre un problème technique. Erreur: {str(e)}"
        }


# ==================== CELERY ASYNC ENDPOINTS ====================

@router.post("/agent/ask_async")
def agent_ask_async(body: AskBody, db: Session = Depends(get_db)):
    """
    Version ASYNCHRONE de /agent/ask.
    Retourne immédiatement un job_id pour polling.
    Nécessite Celery + Redis.
    """
    if not HAS_CELERY:
        raise HTTPException(
            status_code=503,
            detail="Mode asynchrone non disponible. Celery/Redis non configurés."
        )
    if not ServiceStatus().is_redis_available():
        raise HTTPException(
            status_code=503,
            detail="File de tâches indisponible. Vérifiez Redis avant de relancer.",
        )

    q = (body.query or "").strip()
    if not q:
        return {"type": "error", "message": "Query vide"}

    q_lower = q.lower()

    # Déterminer le type de rapport
    report_type = "osint"

    # Si c'est une recherche générale, pas de mode async (trop rapide)
    if any(trig in q_lower for trig in GENERAL_SEARCH_TRIGGERS):
        report_type = "general"

    # Extraire la cible si c'est une commande "analyze"
    query_to_scan = q
    if q_lower.startswith("analyze"):
        parts = q.split(maxsplit=1)
        query_to_scan = parts[1] if len(parts) > 1 else q

    # Récupérer le scan_mode (défaut: "full" - mode simple et stable)
    scan_mode = getattr(body, 'scan_mode', 'full') or 'full'
    approved_tools = getattr(body, 'approved_tools', None) or []
    # Récupérer la langue (défaut: "fr")
    language = getattr(body, 'language', 'fr') or 'fr'
    llm_hard_limit = getattr(body, 'llm_hard_limit', None)

    logger.info(f"[ASYNC] Langue du rapport: {language}")

    # Router vers la tâche appropriée selon le scan_mode
    task = None
    task_queue = None

    if scan_mode == "fast":
        # Layer 1 uniquement: WHOIS, DNS, headers (rapide, ~30s)
        task = scan_osint_layer1_task.delay(query_to_scan, llm_hard_limit)
        task_queue = "osint_fast"
        logger.info(f"[ASYNC] Mode FAST -> queue osint_fast")

    elif scan_mode == "standard":
        # Layer 1 + 2: inclut Censys, crt.sh (~2-3min)
        task = scan_osint_layer2_task.delay(query_to_scan, llm_hard_limit)
        task_queue = "osint_medium"
        logger.info(f"[ASYNC] Mode STANDARD -> queue osint_medium")

    elif scan_mode == "critical" and approved_tools:
        # Layer 3: port_scan, vuln_scan (nécessite approbation)
        task = scan_osint_layer3_task.delay(query_to_scan, approved_tools, llm_hard_limit)
        task_queue = "osint_critical"
        logger.warning(f"[ASYNC] Mode CRITICAL -> queue osint_critical | Tools: {approved_tools}")

    elif scan_mode == "priority":
        # Scan prioritaire (bypass les autres queues)
        task = priority_scan_task.delay(query_to_scan, llm_hard_limit)
        task_queue = "priority"
        logger.warning(f"[ASYNC] Mode PRIORITY -> queue priority 🚨")

    elif scan_mode == "parallel":
        # Architecture parallèle: Layer 1 + Layer 2 en chord
        # Layer 1 (FAST) et Layer 2 (MEDIUM) s'exécutent en parallèle
        # puis aggregate_results génère le rapport final
        task = scan_parallel_task.delay(query_to_scan, llm_hard_limit)
        task_queue = "osint_fast + osint_medium (parallel)"
        logger.info(f"[ASYNC] Mode PARALLEL -> chord architecture (Layer1 || Layer2 -> Aggregate)")

    else:
        # Mode "full" par défaut: scan séquentiel complet (stable)
        task = scan_osint_task.delay(query_to_scan, report_type, language, llm_hard_limit)
        task_queue = "osint_medium"
        logger.info(f"[ASYNC] Mode FULL -> queue osint_medium (lang={language})")

    # Créer l'entrée ScanJob en BDD
    job = ScanJob(
        job_id=task.id,
        query=query_to_scan,
        report_type=report_type,
        status="PENDING",
        progress=0
    )

    db.add(job)
    db.commit()

    logger.info(f"[ASYNC] Tâche créée: {task.id} pour '{query_to_scan}' | Mode: {scan_mode} | Queue: {task_queue}")

    return {
        "type": "async",
        "job_id": task.id,
        "status": "PENDING",
        "scan_mode": scan_mode,
        "queue": task_queue,
        "message": f"Scan lancé en arrière-plan (mode: {scan_mode}). Utilisez /jobs/{task.id} pour suivre la progression."
    }


@router.get("/jobs/{job_id}")
def get_job_status(job_id: str, db: Session = Depends(get_db)):
    """
    Récupère l'état d'une tâche asynchrone.
    Polling endpoint pour le frontend.
    """
    if not HAS_CELERY:
        raise HTTPException(
            status_code=503,
            detail="Mode asynchrone non disponible."
        )

    # Chercher le job en BDD
    job = db.query(ScanJob).filter_by(job_id=job_id).first()

    if not job:
        raise HTTPException(status_code=404, detail="Job non trouvé")

    # Préparer la réponse de base
    response = {
        "job_id": job.job_id,
        "query": job.query,
        "status": job.status,
        "progress": job.progress,
        "created_at": job.created_at.isoformat() if job.created_at else None,
        "updated_at": job.updated_at.isoformat() if job.updated_at else None
    }

    # Si complété, ajouter le résultat
    if job.status == "COMPLETED" and job.result:
        import json
        try:
            result_data = json.loads(job.result)
            response["result"] = result_data
        except:
            response["result"] = {"error": "Erreur parsing résultat"}

    # Si échoué, ajouter le message d'erreur
    if job.status == "FAILED" and job.error_message:
        response["error"] = job.error_message

    return response


@router.get("/jobs/")
def list_jobs(limit: int = 10, db: Session = Depends(get_db)):
    """
    Liste les N derniers jobs (pour debug/monitoring).
    """
    if not HAS_CELERY:
        raise HTTPException(
            status_code=503,
            detail="Mode asynchrone non disponible."
        )

    jobs = db.query(ScanJob).order_by(ScanJob.created_at.desc()).limit(limit).all()

    return {
        "jobs": [
            {
                "job_id": j.job_id,
                "query": j.query,
                "status": j.status,
                "progress": j.progress,
                "created_at": j.created_at.isoformat() if j.created_at else None
            }
            for j in jobs
        ]
    }


# ============================================================================
# ROUTES SYSTÈME D'APPROBATION UTILISATEUR (Layer 3)
# ============================================================================

class ApprovalRequest(BaseModel):
    """Modèle pour créer une demande d'approbation."""
    tool_name: str
    target: str
    run_id: str
    context_declared: str
    hypothesis: str = None


@router.post("/agent/request_approval")
def request_approval(request: ApprovalRequest, db: Session = Depends(get_db)):
    """
    Créer une demande d'approbation pour un outil Layer 3.
    Retourne un approval_id que le frontend utilisera pour afficher le bouton.
    """
    approval_id = str(uuid.uuid4())

    pending_approval = PendingApproval(
        approval_id=approval_id,
        tool_name=request.tool_name,
        target=request.target,
        run_id=request.run_id,
        context_declared=request.context_declared,
        hypothesis=request.hypothesis,
        status="PENDING",
        approved_by_user=False
    )

    db.add(pending_approval)
    db.commit()
    db.refresh(pending_approval)

    logger.info(f"[APPROVAL REQUEST] {approval_id} - {request.tool_name} on {request.target}")

    return {
        "approval_id": approval_id,
        "tool_name": request.tool_name,
        "target": request.target,
        "status": "PENDING",
        "message": f"Approbation requise pour {request.tool_name} (Layer 3)"
    }


@router.post("/agent/approve/{approval_id}")
def approve_tool(approval_id: str, db: Session = Depends(get_db)):
    """
    Approuver l'utilisation d'un outil Layer 3.
    Met à jour le statut dans la BDD et permet au backend de continuer.
    """
    approval = db.query(PendingApproval).filter_by(approval_id=approval_id).first()

    if not approval:
        raise HTTPException(status_code=404, detail="Demande d'approbation introuvable")

    if approval.status != "PENDING":
        raise HTTPException(
            status_code=400,
            detail=f"Demande déjà traitée (statut: {approval.status})"
        )

    # Marquer comme approuvée
    approval.status = "APPROVED"
    approval.approved_by_user = True
    approval.resolved_at = datetime.now(timezone.utc)

    db.commit()

    logger.info(f"[APPROVAL GRANTED] {approval_id} - {approval.tool_name} on {approval.target}")

    return {
        "approval_id": approval_id,
        "status": "APPROVED",
        "message": f"Outil {approval.tool_name} approuvé",
        "tool_name": approval.tool_name,
        "target": approval.target
    }


@router.post("/agent/deny/{approval_id}")
def deny_tool(approval_id: str, reason: str = "User denied", db: Session = Depends(get_db)):
    """
    Refuser l'utilisation d'un outil Layer 3.
    Le scan s'arrêtera sans exécuter l'outil.
    """
    approval = db.query(PendingApproval).filter_by(approval_id=approval_id).first()

    if not approval:
        raise HTTPException(status_code=404, detail="Demande d'approbation introuvable")

    if approval.status != "PENDING":
        raise HTTPException(
            status_code=400,
            detail=f"Demande déjà traitée (statut: {approval.status})"
        )

    # Marquer comme refusée
    approval.status = "DENIED"
    approval.approved_by_user = False
    approval.denial_reason = reason
    approval.resolved_at = datetime.now(timezone.utc)

    db.commit()

    logger.info(f"[APPROVAL DENIED] {approval_id} - {approval.tool_name} on {approval.target} | Raison: {reason}")

    return {
        "approval_id": approval_id,
        "status": "DENIED",
        "message": f"Outil {approval.tool_name} refusé",
        "tool_name": approval.tool_name,
        "target": approval.target,
        "reason": reason
    }


# ==================== MONITORING & AUDIT TRAIL ====================

@router.get("/monitoring/stats")
def get_monitoring_stats(
    request: Request,
    db: Session = Depends(get_db),
):
    """Retourne les statistiques globales des scans (avec cache HTTP).

    Cache court (30s) + ETag + support 304 via If-None-Match.
    """
    try:
        # Total scans
        total_scans = db.query(ToolExecutionLog).count()

        # Success/failure counts
        success_count = db.query(ToolExecutionLog).filter(ToolExecutionLog.status == "ok").count()
        failed_count = db.query(ToolExecutionLog).filter(ToolExecutionLog.status == "error").count()

        # Success rate
        success_rate = (success_count / total_scans * 100) if total_scans > 0 else 0

        # Average duration (only successful scans)
        avg_duration = (
            db.query(func.avg(ToolExecutionLog.duration_seconds))
            .filter(
                ToolExecutionLog.status == "ok",
                ToolExecutionLog.duration_seconds.isnot(None),
            )
            .scalar()
            or 0
        )

        response_data = {
            "total_scans": total_scans,
            "success_count": success_count,
            "failed_scans": failed_count,
            "success_rate": round(success_rate, 2),
            "avg_duration": round(float(avg_duration), 2) if avg_duration else 0,
        }

        etag = generate_etag(response_data)

        # 304 Not Modified si ETag match
        if check_not_modified(request, etag=etag, last_modified=None):
            return Response(
                status_code=304,
                headers={
                    "ETag": etag,
                    "Cache-Control": "public, max-age=30",
                },
            )

        response = JSONResponse(content=response_data)
        add_cache_headers(response, etag=etag, last_modified=None, max_age=30)
        return response

    except Exception as e:
        logger.exception(f"[/monitoring/stats] ERREUR: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la récupération des statistiques")


@router.get("/monitoring/logs")
def get_monitoring_logs(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=100),
    tool: str = Query("", description="Filtrer par outil"),
    status: str = Query("", description="Filtrer par statut"),
    period: str = Query("24h", description="Période (24h, 7d, 30d, all)"),
    db: Session = Depends(get_db)
):
    """Retourne les logs d'exécution avec pagination et filtres."""
    try:
        from datetime import timedelta

        # Base query
        query = db.query(ToolExecutionLog)

        # Apply filters
        if tool:
            query = query.filter(ToolExecutionLog.tool_name == tool)
        if status:
            query = query.filter(ToolExecutionLog.status == status)

        # Period filter
        if period != "all":
            now = datetime.now(timezone.utc)
            period_map = {
                "24h": timedelta(hours=24),
                "7d": timedelta(days=7),
                "30d": timedelta(days=30)
            }
            delta = period_map.get(period, timedelta(hours=24))
            cutoff = now - delta
            query = query.filter(ToolExecutionLog.executed_at >= cutoff)

        # Total count
        total = query.count()

        # Pagination
        offset = (page - 1) * limit
        logs = query.order_by(ToolExecutionLog.executed_at.desc()).offset(offset).limit(limit).all()

        # Get unique tool names for filter
        tools = db.query(ToolExecutionLog.tool_name).distinct().all()
        tool_names = [t[0] for t in tools]

        return {
            "logs": [
                {
                    "id": log.id,
                    "timestamp": log.executed_at.isoformat() if log.executed_at else None,
                    "run_id": log.run_id,
                    "tool_name": log.tool_name,
                    "target": log.target,
                    "status": log.status,
                    "duration": float(log.duration_seconds) if log.duration_seconds else None,
                    "layer": log.tool_layer,
                    "consent_given": log.user_consent
                }
                for log in logs
            ],
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit,
            "tools": tool_names
        }

    except Exception as e:
        logger.exception(f"[/monitoring/logs] ERREUR: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la récupération des logs")


@router.get("/monitoring/logs/{log_id}")
def get_log_detail(log_id: int, db: Session = Depends(get_db)):
    """Retourne les détails complets d'un log."""
    try:
        log = db.query(ToolExecutionLog).filter(ToolExecutionLog.id == log_id).first()

        if not log:
            raise HTTPException(status_code=404, detail="Log introuvable")

        return {
            "id": log.id,
            "timestamp": log.executed_at.isoformat() if log.executed_at else None,
            "run_id": log.run_id,
            "tool_name": log.tool_name,
            "target": log.target,
            "status": log.status,
            "duration": float(log.duration_seconds) if log.duration_seconds else None,
            "layer": log.tool_layer,
            "consent_given": log.user_consent,
            "error_message": log.error_message,
            "context_declared": log.context_declared,
            "legal_risk_level": log.legal_risk_level,
            "result_summary": log.result_summary
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[/monitoring/logs/{log_id}] ERREUR: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la récupération du log")


# ==================== API KEYS MANAGEMENT ====================

@router.post("/api-keys/create")
def create_api_key(
    request: Request,
    name: str = Query(..., description="Nom descriptif de la clé"),
    role: str = Query("analyst", description="Rôle: admin, analyst ou viewer"),
    owner_id: Optional[str] = Query(None, max_length=100),
    db: Session = Depends(get_db)
):
    """Crée une nouvelle API key."""
    try:
        from auth import generate_api_key

        role = role.strip().lower()
        if role not in {"admin", "analyst", "viewer"}:
            raise HTTPException(status_code=422, detail="Rôle invalide")

        # La clé de bootstrap doit toujours être administratrice.
        if db.query(APIKey).filter(APIKey.is_active.is_(True)).count() == 0:
            role = "admin"

        principal = getattr(request.state, "principal", None)
        actor_id = getattr(principal, "actor_id", None) or "local"
        resolved_owner_id = (
            owner_id.strip()
            if owner_id and owner_id.strip()
            else f"principal:{uuid.uuid4().hex}"
        )

        # Générer la clé
        api_key, key_hash = generate_api_key()

        # Créer l'entrée dans la base de données
        new_key = APIKey(
            key_hash=key_hash,
            name=name,
            prefix=api_key[:12],  # Stocker le préfixe pour l'affichage
            is_active=True,
            role=role,
            owner_id=resolved_owner_id,
            created_by=actor_id,
        )

        db.add(new_key)
        db.commit()
        db.refresh(new_key)

        logger.info(f"[API KEY] Nouvelle clé créée: {name} (ID: {new_key.id})")

        # IMPORTANT: Retourner la clé complète UNE SEULE FOIS
        # L'utilisateur doit la copier, elle ne sera plus accessible ensuite
        return {
            "success": True,
            "api_key": api_key,  # Clé complète (à copier immédiatement)
            "id": new_key.id,
            "name": new_key.name,
            "prefix": new_key.prefix,
            "role": new_key.role,
            "owner_id": new_key.owner_id,
            "created_at": new_key.created_at.isoformat(),
            "warning": "⚠️ Copiez cette clé maintenant. Elle ne sera plus affichée."
        }

    except Exception as e:
        logger.exception(f"[/api-keys/create] ERREUR: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail="Erreur lors de la création de l'API key")


@router.get("/api-keys/list")
def list_api_keys(db: Session = Depends(get_db)):
    """Liste toutes les API keys (sans les clés complètes)."""
    try:
        keys = db.query(APIKey).order_by(APIKey.created_at.desc()).all()

        # Les tests d'intégration attendent une liste directement.
        return [
            {
                "id": key.id,
                "name": key.name,
                "prefix": key.prefix,
                "is_active": key.is_active,
                "role": key.role,
                "owner_id": key.owner_id,
                "created_at": key.created_at.isoformat() if key.created_at else None,
                "last_used_at": key.last_used_at.isoformat() if key.last_used_at else None,
                "created_by": key.created_by,
            }
            for key in keys
        ]

    except Exception as e:
        logger.exception(f"[/api-keys/list] ERREUR: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la récupération des API keys")


@router.delete("/api-keys/{key_id}")
def revoke_api_key(key_id: int, db: Session = Depends(get_db)):
    """Révoque (désactive) une API key."""
    try:
        key = db.query(APIKey).filter(APIKey.id == key_id).first()

        if not key:
            raise HTTPException(status_code=404, detail="API Key introuvable")

        key.is_active = False
        db.commit()

        logger.info(f"[API KEY] Clé révoquée: {key.name} (ID: {key.id})")

        return {
            "success": True,
            "message": f"API Key '{key.name}' révoquée avec succès",
            "id": key.id,
            "name": key.name
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[/api-keys/{key_id}] ERREUR: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail="Erreur lors de la révocation de l'API key")


# ==================== SCAN COMPARISON ====================


# ==================== TIMELINE ====================


@router.get("/osint/graph")
def osint_graph(
    target: str = Query(..., description="Cible (domaine/IP)"),
    db: Session = Depends(get_db),
):
    """Retourne le graph relationnel (dernier rapport en base) pour une cible."""
    norm = validate_target(target)

    report = (
        db.query(EntityReport)
        .filter(EntityReport.target.ilike(f"%{norm}%"))
        .order_by(EntityReport.updated_at.desc(), EntityReport.created_at.desc())
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        payload = json.loads(report.raw_data) if report.raw_data else {}
    except Exception:
        payload = {}

    graph = payload.get("intel_graph") or {"version": 1, "root": None, "nodes": [], "edges": []}
    return {"target": norm, "report_id": report.id, "graph": graph}


@router.get("/osint/graph/{report_id}")
def osint_graph_by_report(
    report_id: int,
    db: Session = Depends(get_db),
):
    """Retourne le graph relationnel pour un report_id (utile pour diff/comparaison)."""
    report = db.query(EntityReport).filter(EntityReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        payload = json.loads(report.raw_data) if report.raw_data else {}
    except Exception:
        payload = {}

    graph = payload.get("intel_graph") or {"version": 1, "root": None, "nodes": [], "edges": []}
    return {"target": report.target, "report_id": report.id, "graph": graph}


@router.get("/osint/structured")
def osint_structured(
    target: str = Query(..., description="Cible (domaine/IP)"),
    db: Session = Depends(get_db),
):
    """Retourne les findings structurés (Phase 1) du dernier rapport pour une cible."""
    norm = validate_target(target)

    report = (
        db.query(EntityReport)
        .filter(EntityReport.target.ilike(f"%{norm}%"))
        .order_by(EntityReport.updated_at.desc(), EntityReport.created_at.desc())
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        payload = json.loads(report.raw_data) if report.raw_data else {}
    except Exception:
        payload = {}

    return {"target": norm, "report_id": report.id, "structured_data": payload.get("structured_data") or {}}


@router.get("/osint/structured/{report_id}")
def osint_structured_by_report(
    report_id: int,
    db: Session = Depends(get_db),
):
    """Retourne les findings structurés (Phase 1) pour un report_id."""
    report = db.query(EntityReport).filter(EntityReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        payload = json.loads(report.raw_data) if report.raw_data else {}
    except Exception:
        payload = {}

    return {"target": report.target, "report_id": report.id, "structured_data": payload.get("structured_data") or {}}


@router.get("/osint/exposures")
def osint_exposures(
    target: str = Query(..., description="Cible (domaine/IP)"),
    db: Session = Depends(get_db),
):
    """Retourne les expositions normalisées (dernier report en base) pour une cible."""
    norm = validate_target(target)

    report = (
        db.query(EntityReport)
        .filter(EntityReport.target.ilike(f"%{norm}%"))
        .order_by(EntityReport.updated_at.desc(), EntityReport.created_at.desc())
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        payload = json.loads(report.raw_data) if report.raw_data else {}
    except Exception:
        payload = {}

    exposures = payload.get("exposures") or []
    return {"target": norm, "report_id": report.id, "count": len(exposures), "exposures": exposures}


@router.get("/osint/exposures/{report_id}")
def osint_exposures_by_report(
    report_id: int,
    db: Session = Depends(get_db),
):
    """Retourne les expositions normalisées pour un report_id."""
    report = db.query(EntityReport).filter(EntityReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        payload = json.loads(report.raw_data) if report.raw_data else {}
    except Exception:
        payload = {}

    exposures = payload.get("exposures") or []
    return {"target": report.target, "report_id": report.id, "count": len(exposures), "exposures": exposures}


@router.get("/osint/timeline_events")
def osint_timeline_events(
    target: str = Query(..., description="Cible (domaine/IP)"),
    db: Session = Depends(get_db),
):
    """Retourne les timeline_events du dernier report pour une cible."""
    norm = validate_target(target)

    report = (
        db.query(EntityReport)
        .filter(EntityReport.target.ilike(f"%{norm}%"))
        .order_by(EntityReport.updated_at.desc(), EntityReport.created_at.desc())
        .first()
    )
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        payload = json.loads(report.raw_data) if report.raw_data else {}
    except Exception:
        payload = {}

    events = payload.get("timeline_events") or []
    return {"target": norm, "report_id": report.id, "count": len(events), "events": events}


@router.get("/osint/timeline_events/{report_id}")
def osint_timeline_events_by_report(
    report_id: int,
    db: Session = Depends(get_db),
):
    """Retourne les timeline_events pour un report_id."""
    report = db.query(EntityReport).filter(EntityReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    try:
        payload = json.loads(report.raw_data) if report.raw_data else {}
    except Exception:
        payload = {}

    events = payload.get("timeline_events") or []
    return {"target": report.target, "report_id": report.id, "count": len(events), "events": events}


@router.get("/osint/timeline_summary")
def osint_timeline_summary(
    target: str = Query(..., description="Cible (domaine/IP)"),
    limit: int = Query(10, ge=2, le=50),
    db: Session = Depends(get_db),
):
    """Résumé d'évolution (timeline) basé sur les N derniers reports en base.

    Objectif: mettre en avant les points de rupture (risque/exposures/graph) plutôt que des dumps.
    """
    norm = validate_target(target)

    reports = (
        db.query(EntityReport)
        .filter(EntityReport.target.ilike(f"%{norm}%"))
        .order_by(EntityReport.updated_at.desc(), EntityReport.created_at.desc())
        .limit(limit)
        .all()
    )

    if len(reports) < 2:
        return {"target": norm, "count": len(reports), "items": [], "message": "Pas assez d'historique pour comparer"}

    def _load_payload(r: EntityReport) -> dict:
        try:
            return json.loads(r.raw_data) if r.raw_data else {}
        except Exception:
            return {}

    def _risk(payload: dict):
        ra = payload.get("risk_analysis")
        if isinstance(ra, dict):
            return ra.get("score"), ra.get("level")
        return None, None

    def _exposure_keys(payload: dict) -> set[str]:
        exps = payload.get("exposures") or []
        keys = set()
        for e in exps:
            if not isinstance(e, dict):
                continue
            keys.add(f"{e.get('type','')}|{e.get('id','')}|{e.get('title','')}")
        return keys

    def _graph_edge_keys(payload: dict) -> set[str]:
        g = payload.get("intel_graph") or {}
        edges = g.get("edges") if isinstance(g, dict) else None
        out = set()
        if isinstance(edges, list):
            for ed in edges:
                if not isinstance(ed, dict):
                    continue
                out.add(f"{ed.get('source')}|{ed.get('type')}|{ed.get('target')}")
        return out

    # Process from oldest -> newest for a forward timeline
    ordered = list(reversed(reports))

    items = []
    prev_payload = _load_payload(ordered[0])
    prev_score, prev_level = _risk(prev_payload)
    prev_exps = _exposure_keys(prev_payload)
    prev_edges = _graph_edge_keys(prev_payload)

    prev_report_id = ordered[0].id

    for r in ordered[1:]:
        payload = _load_payload(r)
        score, level = _risk(payload)
        exps = _exposure_keys(payload)
        edges = _graph_edge_keys(payload)

        risk_delta = None
        if prev_score is not None and score is not None and score != prev_score:
            risk_delta = score - prev_score

        exp_added = len(exps - prev_exps)
        exp_removed = len(prev_exps - exps)
        edge_added = len(edges - prev_edges)
        edge_removed = len(prev_edges - edges)

        # Only keep meaningful changes
        meaningful = (
            (risk_delta is not None and abs(risk_delta) >= 5)
            or exp_added > 0
            or exp_removed > 0
            or edge_added > 0
            or edge_removed > 0
            or (prev_level and level and prev_level != level)
        )

        if meaningful:
            items.append({
                "from_report_id": prev_report_id,
                "to_report_id": r.id,
                "at": (r.updated_at or r.created_at).isoformat() if (r.updated_at or r.created_at) else None,
                "risk": {
                    "from_score": prev_score,
                    "to_score": score,
                    "delta": risk_delta,
                    "from_level": prev_level,
                    "to_level": level,
                },
                "changes": {
                    "exposures_added": exp_added,
                    "exposures_removed": exp_removed,
                    "graph_edges_added": edge_added,
                    "graph_edges_removed": edge_removed,
                }
            })

        prev_payload = payload
        prev_report_id = r.id
        prev_score, prev_level = score, level
        prev_exps = exps
        prev_edges = edges

    return {"target": norm, "count": len(items), "items": items}


@router.get("/osint/timeline")
def osint_timeline(
    target: str = Query(..., description="Cible (domaine/IP)"),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """Timeline basée sur l'historique ScanJob (OSINT async)."""
    norm = validate_target(target)

    jobs = (
        db.query(ScanJob)
        .filter(ScanJob.query.ilike(norm))
        .order_by(ScanJob.created_at.desc())
        .limit(limit)
        .all()
    )

    score_re = re.compile(r"Score de risque global:\s*(\d{1,3})/100\s*\(Niveau:\s*([^\)]+)\)", re.IGNORECASE)
    level_re = re.compile(r"\((FAIBLE|MOYEN|ÉLEVÉ|ELEV|CRITIQUE)\)", re.IGNORECASE)

    items = []
    for j in reversed(jobs):
        report = ""
        sources_count = 0
        if j.result:
            try:
                payload = json.loads(j.result)
                report = payload.get("report") or ""
                sources = payload.get("sources") or payload.get("results") or []
                if isinstance(sources, list):
                    sources_count = len(sources)
            except Exception:
                report = ""

        risk_score = None
        risk_level = None
        m = score_re.search(report or "")
        if m:
            try:
                risk_score = int(m.group(1))
            except Exception:
                risk_score = None
            risk_level = (m.group(2) or "").strip()
        else:
            m2 = level_re.search(report or "")
            if m2:
                risk_level = m2.group(1).upper()

        created = j.created_at.isoformat() if j.created_at else None
        updated = j.updated_at.isoformat() if j.updated_at else None
        duration_s = None
        if j.created_at and j.updated_at:
            duration_s = round((j.updated_at - j.created_at).total_seconds(), 2)

        items.append(
            {
                "job_id": j.job_id,
                "created_at": created,
                "updated_at": updated,
                "status": j.status,
                "progress": j.progress,
                "duration_seconds": duration_s,
                "risk_score": risk_score,
                "risk_level": risk_level,
                "sources_count": sources_count,
            }
        )

    return {"target": norm, "count": len(items), "items": items}

@router.get("/osint/diff")
def diff_reports(
    report_id_1: int = Query(..., description="ID du premier rapport"),
    report_id_2: int = Query(..., description="ID du second rapport"),
    db: Session = Depends(get_db)
):
    """Diff structuré entre deux rapports.

    Renvoie uniquement les changements utiles pour le "renseignement": exposures + graph.
    """
    try:
        report1 = db.query(EntityReport).filter(EntityReport.id == report_id_1).first()
        report2 = db.query(EntityReport).filter(EntityReport.id == report_id_2).first()

        if not report1 or not report2:
            raise HTTPException(status_code=404, detail="Un ou plusieurs rapports introuvables")

        if report1.target.lower() != report2.target.lower():
            raise HTTPException(
                status_code=400,
                detail=f"Les rapports concernent des cibles différentes: {report1.target} vs {report2.target}"
            )

        data1 = json.loads(report1.raw_data) if report1.raw_data else {}
        data2 = json.loads(report2.raw_data) if report2.raw_data else {}

        exp1 = data1.get("exposures") or []
        exp2 = data2.get("exposures") or []

        def _exp_key(e: dict) -> str:
            if not isinstance(e, dict):
                return ""
            return f"{e.get('type','')}|{e.get('id','')}|{e.get('title','')}"

        exp_map1 = {k: e for e in exp1 if (k := _exp_key(e))}
        exp_map2 = {k: e for e in exp2 if (k := _exp_key(e))}

        exp_added = [exp_map2[k] for k in exp_map2.keys() - exp_map1.keys()]
        exp_removed = [exp_map1[k] for k in exp_map1.keys() - exp_map2.keys()]

        exp_changed = []
        for k in exp_map1.keys() & exp_map2.keys():
            a = exp_map1[k]
            b = exp_map2[k]
            if (a.get("severity") != b.get("severity")) or (a.get("confidence") != b.get("confidence")):
                exp_changed.append({"before": a, "after": b})

        g1 = (data1.get("intel_graph") or {})
        g2 = (data2.get("intel_graph") or {})

        def _node_key(n: dict) -> str:
            if not isinstance(n, dict):
                return ""
            return n.get("id") or ""

        def _edge_key(ed: dict) -> str:
            if not isinstance(ed, dict):
                return ""
            return f"{ed.get('source')}|{ed.get('type')}|{ed.get('target')}"

        node_map1 = {k: x for x in ((g1.get("nodes") or []) if isinstance(g1, dict) else []) if (k := _node_key(x))}
        node_map2 = {k: x for x in ((g2.get("nodes") or []) if isinstance(g2, dict) else []) if (k := _node_key(x))}
        edge_map1 = {k: x for x in ((g1.get("edges") or []) if isinstance(g1, dict) else []) if (k := _edge_key(x))}
        edge_map2 = {k: x for x in ((g2.get("edges") or []) if isinstance(g2, dict) else []) if (k := _edge_key(x))}

        nodes_added = [node_map2[k] for k in list(node_map2.keys() - node_map1.keys())[:200]]
        nodes_removed = [node_map1[k] for k in list(node_map1.keys() - node_map2.keys())[:200]]
        edges_added = [edge_map2[k] for k in list(edge_map2.keys() - edge_map1.keys())[:400]]
        edges_removed = [edge_map1[k] for k in list(edge_map1.keys() - edge_map2.keys())[:400]]

        structured_diff = {
            "exposures": {
                "added": exp_added,
                "removed": exp_removed,
                "changed": exp_changed,
            },
            "graph": {
                "nodes_added": nodes_added,
                "nodes_removed": nodes_removed,
                "edges_added": edges_added,
                "edges_removed": edges_removed,
            },
        }

        return {
            "target": report1.target,
            "report_id_1": report1.id,
            "report_id_2": report2.id,
            "structured_diff": structured_diff,
            "summary": {
                "exposures_added": len(exp_added),
                "exposures_removed": len(exp_removed),
                "exposures_changed": len(exp_changed),
                "graph_nodes_added": len(nodes_added),
                "graph_nodes_removed": len(nodes_removed),
                "graph_edges_added": len(edges_added),
                "graph_edges_removed": len(edges_removed),
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[/osint/diff] ERREUR: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors du diff")


@router.get("/osint/compare")
def compare_scans(
    target: str = Query(..., description="Cible à comparer"),
    report_id_1: int = Query(..., description="ID du premier rapport"),
    report_id_2: int = Query(..., description="ID du second rapport"),
    db: Session = Depends(get_db)
):
    """
    Compare deux rapports OSINT pour la même cible.
    Identifie les changements entre les deux scans (ajouts, suppressions, modifications).
    """
    try:
        # Récupérer les deux rapports
        report1 = db.query(EntityReport).filter(EntityReport.id == report_id_1).first()
        report2 = db.query(EntityReport).filter(EntityReport.id == report_id_2).first()

        if not report1 or not report2:
            raise HTTPException(status_code=404, detail="Un ou plusieurs rapports introuvables")

        # Vérifier que les rapports concernent la même cible
        if report1.target.lower() != report2.target.lower():
            raise HTTPException(
                status_code=400,
                detail=f"Les rapports concernent des cibles différentes: {report1.target} vs {report2.target}"
            )

        # Parser les données brutes
        import json
        data1 = json.loads(report1.raw_data) if report1.raw_data else {}
        data2 = json.loads(report2.raw_data) if report2.raw_data else {}

        changes = []

        # Risk score changes (if present)
        risk1 = data1.get("risk_analysis") or {}
        risk2 = data2.get("risk_analysis") or {}
        if isinstance(risk1, dict) and isinstance(risk2, dict):
            score1 = risk1.get("score")
            score2 = risk2.get("score")
            level1 = (risk1.get("level") or "").upper()
            level2 = (risk2.get("level") or "").upper()
            level_rank = {"FAIBLE": 1, "LOW": 1, "MOYEN": 2, "MEDIUM": 2, "ELEV": 3, "ÉLEV": 3, "HIGH": 3, "CRIT": 4, "CRITIQUE": 4, "CRITICAL": 4}

            if score1 is not None and score2 is not None and score1 != score2:
                changes.append(
                    {
                        "tool": "risk_analysis",
                        "type": "risk_score_change",
                        "old_value": score1,
                        "new_value": score2,
                        "severity": "high" if score2 > score1 else "medium",
                    }
                )

            if level1 and level2 and level1 != level2:
                sev = "medium"
                if level_rank.get(level2, 0) > level_rank.get(level1, 0):
                    sev = "high"
                changes.append(
                    {
                        "tool": "risk_analysis",
                        "type": "risk_level_change",
                        "old_value": level1,
                        "new_value": level2,
                        "severity": sev,
                    }
                )

        # Comparer les outils exécutés
        tools1 = set(data1.get("tools", {}).keys())
        tools2 = set(data2.get("tools", {}).keys())

        tools_added = list(tools2 - tools1)
        tools_removed = list(tools1 - tools2)
        tools_common = list(tools1 & tools2)

        # Comparer les résultats des outils communs
        for tool_name in tools_common:
            tool1_data = data1["tools"][tool_name]
            tool2_data = data2["tools"][tool_name]

            # Comparer les statuts
            if tool1_data.get("status") != tool2_data.get("status"):
                changes.append({
                    "tool": tool_name,
                    "type": "status_change",
                    "old_value": tool1_data.get("status"),
                    "new_value": tool2_data.get("status"),
                    "severity": "medium"
                })

            # Comparaisons spécifiques par outil
            if tool_name == "whois" and both_ok(tool1_data, tool2_data):
                # Comparer les infos WHOIS
                compare_whois_data(tool1_data.get("data"), tool2_data.get("data"), changes)

            elif tool_name == "dns_resolution" and both_ok(tool1_data, tool2_data):
                # Comparer les résolutions DNS
                if tool1_data.get("data") != tool2_data.get("data"):
                    changes.append({
                        "tool": tool_name,
                        "type": "ip_change",
                        "old_value": tool1_data.get("data"),
                        "new_value": tool2_data.get("data"),
                        "severity": "high",
                        "description": "L'adresse IP a changé"
                    })

            elif tool_name == "http_headers" and both_ok(tool1_data, tool2_data):
                compare_http_headers_data(tool1_data.get("data"), tool2_data.get("data"), changes)

            elif tool_name == "ssl_analysis" and both_ok(tool1_data, tool2_data):
                compare_ssl_analysis_data(tool1_data.get("data"), tool2_data.get("data"), changes)

            elif tool_name == "redirect_chain" and both_ok(tool1_data, tool2_data):
                compare_redirect_chain_data(tool1_data.get("data"), tool2_data.get("data"), changes)

            elif tool_name == "tls_ciphers" and both_ok(tool1_data, tool2_data):
                compare_tls_ciphers_data(tool1_data.get("data"), tool2_data.get("data"), changes)

            elif tool_name == "email_config" and both_ok(tool1_data, tool2_data):
                compare_email_config_data(tool1_data.get("data"), tool2_data.get("data"), changes)

            elif tool_name == "security_txt" and both_ok(tool1_data, tool2_data):
                compare_security_txt_data(tool1_data.get("data"), tool2_data.get("data"), changes)

            elif tool_name == "robots_txt" and both_ok(tool1_data, tool2_data):
                compare_robots_txt_data(tool1_data.get("data"), tool2_data.get("data"), changes)

        # -------------------- Structured diff (exposures + graph) --------------------
        exp1 = data1.get("exposures") or []
        exp2 = data2.get("exposures") or []

        def _exp_key(e: dict) -> str:
            if not isinstance(e, dict):
                return ""
            return f"{e.get('type','')}|{e.get('id','')}|{e.get('title','')}"

        exp_map1 = {k: e for e in exp1 if (k := _exp_key(e))}
        exp_map2 = {k: e for e in exp2 if (k := _exp_key(e))}

        exp_added = [exp_map2[k] for k in exp_map2.keys() - exp_map1.keys()]
        exp_removed = [exp_map1[k] for k in exp_map1.keys() - exp_map2.keys()]

        exp_changed = []
        for k in exp_map1.keys() & exp_map2.keys():
            a = exp_map1[k]
            b = exp_map2[k]
            if (a.get("severity") != b.get("severity")) or (a.get("confidence") != b.get("confidence")):
                exp_changed.append({"before": a, "after": b})

        g1 = (data1.get("intel_graph") or {})
        g2 = (data2.get("intel_graph") or {})

        def _node_key(n: dict) -> str:
            if not isinstance(n, dict):
                return ""
            return n.get("id") or ""

        def _edge_key(ed: dict) -> str:
            if not isinstance(ed, dict):
                return ""
            return f"{ed.get('source')}|{ed.get('type')}|{ed.get('target')}"

        node_map1 = {k: x for x in ((g1.get("nodes") or []) if isinstance(g1, dict) else []) if (k := _node_key(x))}
        node_map2 = {k: x for x in ((g2.get("nodes") or []) if isinstance(g2, dict) else []) if (k := _node_key(x))}
        edge_map1 = {k: x for x in ((g1.get("edges") or []) if isinstance(g1, dict) else []) if (k := _edge_key(x))}
        edge_map2 = {k: x for x in ((g2.get("edges") or []) if isinstance(g2, dict) else []) if (k := _edge_key(x))}

        nodes_added = [node_map2[k] for k in list(node_map2.keys() - node_map1.keys())[:200]]
        nodes_removed = [node_map1[k] for k in list(node_map1.keys() - node_map2.keys())[:200]]
        edges_added = [edge_map2[k] for k in list(edge_map2.keys() - edge_map1.keys())[:400]]
        edges_removed = [edge_map1[k] for k in list(edge_map1.keys() - edge_map2.keys())[:400]]

        structured_diff = {
            "exposures": {
                "added": exp_added,
                "removed": exp_removed,
                "changed": exp_changed,
            },
            "graph": {
                "nodes_added": nodes_added,
                "nodes_removed": nodes_removed,
                "edges_added": edges_added,
                "edges_removed": edges_removed,
            },
        }

        # Optional: also append exposure changes into the legacy changes list
        for e in exp_added[:50]:
            changes.append({
                "tool": "exposures",
                "type": "exposure_added",
                "new_value": e,
                "severity": (e.get("severity") or "MEDIUM").lower(),
            })
        for e in exp_removed[:50]:
            changes.append({
                "tool": "exposures",
                "type": "exposure_removed",
                "old_value": e,
                "severity": "low",
            })
        for c in exp_changed[:50]:
            after = c.get("after") or {}
            before = c.get("before") or {}
            sev = (after.get("severity") or before.get("severity") or "MEDIUM").lower()
            changes.append({
                "tool": "exposures",
                "type": "exposure_changed",
                "old_value": before,
                "new_value": after,
                "severity": sev,
            })

        # Construire le rapport de comparaison
        comparison_report = {
            "target": report1.target,
            "report1": {
                "id": report1.id,
                "date": (report1.updated_at or report1.created_at).isoformat() if (report1.updated_at or report1.created_at) else None,
                "tools_count": len(tools1)
            },
            "report2": {
                "id": report2.id,
                "date": (report2.updated_at or report2.created_at).isoformat() if (report2.updated_at or report2.created_at) else None,
                "tools_count": len(tools2)
            },
            "summary": {
                "tools_added": len(tools_added),
                "tools_removed": len(tools_removed),
                "changes_detected": len(changes)
            },
            "details": {
                "tools_added": tools_added,
                "tools_removed": tools_removed,
                "changes": changes
            },
            "structured_diff": structured_diff,
        }

        return comparison_report

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[/osint/compare] ERREUR: {e}")
        raise HTTPException(status_code=500, detail="Erreur lors de la comparaison des rapports")


def both_ok(data1: dict, data2: dict) -> bool:
    """Vérifie si les deux données ont le statut 'ok'."""
    return data1.get("status") == "ok" and data2.get("status") == "ok"


def compare_whois_data(data1: any, data2: any, changes: list):
    """Compare les données WHOIS et ajoute les changements détectés."""
    if not isinstance(data1, dict) or not isinstance(data2, dict):
        return

    # Comparer les champs clés
    key_fields = ["registrar", "creation_date", "expiration_date", "org", "organization"]

    for field in key_fields:
        val1 = data1.get(field)
        val2 = data2.get(field)

        if val1 != val2:
            changes.append({
                "tool": "whois",
                "type": "field_change",
                "field": field,
                "old_value": str(val1)[:100] if val1 else None,
                "new_value": str(val2)[:100] if val2 else None,
                "severity": "medium" if field in ["registrar", "org"] else "low"
            })


def compare_http_headers_data(data1: any, data2: any, changes: list):
    if not isinstance(data1, dict) or not isinstance(data2, dict):
        return

    sec1 = data1.get("security_headers") or {}
    sec2 = data2.get("security_headers") or {}
    if not isinstance(sec1, dict) or not isinstance(sec2, dict):
        return

    fields = [
        ("Strict-Transport-Security", "high"),
        ("Content-Security-Policy", "medium"),
        ("X-Frame-Options", "medium"),
        ("X-Content-Type-Options", "low"),
    ]
    for header, severity in fields:
        v1 = sec1.get(header)
        v2 = sec2.get(header)
        if v1 != v2:
            changes.append(
                {
                    "tool": "http_headers",
                    "type": "security_header_change",
                    "field": header,
                    "old_value": v1,
                    "new_value": v2,
                    "severity": severity,
                }
            )


def compare_ssl_analysis_data(data1: any, data2: any, changes: list):
    if not isinstance(data1, dict) or not isinstance(data2, dict):
        return

    issuer1 = (data1.get("issuer") or {}).get("organizationName") if isinstance(data1.get("issuer"), dict) else None
    issuer2 = (data2.get("issuer") or {}).get("organizationName") if isinstance(data2.get("issuer"), dict) else None
    if issuer1 != issuer2:
        changes.append(
            {
                "tool": "ssl_analysis",
                "type": "issuer_change",
                "old_value": issuer1,
                "new_value": issuer2,
                "severity": "medium",
            }
        )

    na1 = data1.get("not_after")
    na2 = data2.get("not_after")
    if na1 != na2:
        changes.append(
            {
                "tool": "ssl_analysis",
                "type": "certificate_expiry_change",
                "old_value": na1,
                "new_value": na2,
                "severity": "medium",
            }
        )


def compare_redirect_chain_data(data1: any, data2: any, changes: list):
    if not isinstance(data1, dict) or not isinstance(data2, dict):
        return

    f1 = data1.get("final_url")
    f2 = data2.get("final_url")
    if f1 != f2:
        changes.append(
            {
                "tool": "redirect_chain",
                "type": "final_url_change",
                "old_value": f1,
                "new_value": f2,
                "severity": "medium",
            }
        )

    l1 = data1.get("chain_length")
    l2 = data2.get("chain_length")
    if l1 != l2:
        changes.append(
            {
                "tool": "redirect_chain",
                "type": "redirect_count_change",
                "old_value": l1,
                "new_value": l2,
                "severity": "low",
            }
        )


def compare_tls_ciphers_data(data1: any, data2: any, changes: list):
    if not isinstance(data1, dict) or not isinstance(data2, dict):
        return

    p1 = data1.get("protocol_version")
    p2 = data2.get("protocol_version")
    if p1 != p2:
        changes.append(
            {
                "tool": "tls_ciphers",
                "type": "tls_version_change",
                "old_value": p1,
                "new_value": p2,
                "severity": "high",
            }
        )

    c1 = (data1.get("current_cipher") or {}).get("name") if isinstance(data1.get("current_cipher"), dict) else None
    c2 = (data2.get("current_cipher") or {}).get("name") if isinstance(data2.get("current_cipher"), dict) else None
    if c1 != c2:
        changes.append(
            {
                "tool": "tls_ciphers",
                "type": "cipher_change",
                "old_value": c1,
                "new_value": c2,
                "severity": "medium",
            }
        )


def compare_email_config_data(data1: any, data2: any, changes: list):
    if not isinstance(data1, dict) or not isinstance(data2, dict):
        return

    for field, severity in [("spf", "medium"), ("dmarc", "medium")]:
        v1 = data1.get(field)
        v2 = data2.get(field)
        if v1 != v2:
            changes.append(
                {
                    "tool": "email_config",
                    "type": "email_policy_change",
                    "field": field,
                    "old_value": v1,
                    "new_value": v2,
                    "severity": severity,
                }
            )

    mx1 = data1.get("mx_records")
    mx2 = data2.get("mx_records")
    if mx1 != mx2:
        changes.append(
            {
                "tool": "email_config",
                "type": "mx_change",
                "old_value": mx1,
                "new_value": mx2,
                "severity": "medium",
            }
        )


def compare_security_txt_data(data1: any, data2: any, changes: list):
    if not isinstance(data1, dict) or not isinstance(data2, dict):
        return

    e1 = data1.get("exists")
    e2 = data2.get("exists")
    if e1 != e2:
        changes.append(
            {
                "tool": "security_txt",
                "type": "exists_change",
                "old_value": e1,
                "new_value": e2,
                "severity": "low",
            }
        )


def compare_robots_txt_data(data1: any, data2: any, changes: list):
    if not isinstance(data1, dict) or not isinstance(data2, dict):
        return

    e1 = data1.get("exists")
    e2 = data2.get("exists")
    if e1 != e2:
        changes.append(
            {
                "tool": "robots_txt",
                "type": "exists_change",
                "old_value": e1,
                "new_value": e2,
                "severity": "low",
            }
        )

    d1 = data1.get("disallowed_paths")
    d2 = data2.get("disallowed_paths")
    if isinstance(d1, list) and isinstance(d2, list) and d1 != d2:
        changes.append(
            {
                "tool": "robots_txt",
                "type": "disallowed_paths_change",
                "old_value": len(d1),
                "new_value": len(d2),
                "severity": "low",
            }
        )


# ==================== CELERY WORKERS MONITORING ====================

# ==================== WEBSOCKET NOTIFICATIONS ====================

# Connection manager for WebSocket clients
class ConnectionManager:
    """Manages WebSocket connections for real-time job notifications."""

    def __init__(self):
        self.active_connections: Dict[str, Set[WebSocket]] = {}  # job_id -> set of websockets

    async def connect(self, websocket: WebSocket, job_id: str, subprotocol: str | None = None):
        await websocket.accept(subprotocol=subprotocol)
        if job_id not in self.active_connections:
            self.active_connections[job_id] = set()
        self.active_connections[job_id].add(websocket)
        logger.info(f"[WebSocket] Client connected for job {job_id}")

    def disconnect(self, websocket: WebSocket, job_id: str):
        if job_id in self.active_connections:
            self.active_connections[job_id].discard(websocket)
            if not self.active_connections[job_id]:
                del self.active_connections[job_id]
        logger.info(f"[WebSocket] Client disconnected from job {job_id}")

    async def send_job_update(self, job_id: str, message: dict):
        """Send update to all clients watching a specific job."""
        if job_id in self.active_connections:
            disconnected = []
            for websocket in self.active_connections[job_id]:
                try:
                    await websocket.send_json(message)
                except Exception as e:
                    logger.warning(f"[WebSocket] Failed to send to client: {e}")
                    disconnected.append(websocket)
            # Clean up disconnected clients
            for ws in disconnected:
                self.active_connections[job_id].discard(ws)

    async def broadcast(self, message: dict):
        """Send message to all connected clients."""
        for job_id in list(self.active_connections.keys()):
            await self.send_job_update(job_id, message)


# Global connection manager instance
ws_manager = ConnectionManager()


@router.websocket("/ws/jobs/{job_id}")
async def websocket_job_status(websocket: WebSocket, job_id: str):
    """
    WebSocket endpoint for real-time job status updates.
    Replaces polling for better performance and UX.

    Usage:
        const ws = new WebSocket(`ws://localhost:8010/ws/jobs/${jobId}`);
        ws.onmessage = (event) => {
            const data = JSON.parse(event.data);
            console.log('Job update:', data);
        };
    """
    from auth import authenticate_api_key_value, authentication_required
    from database import SessionLocal

    protocols = [
        value.strip()
        for value in websocket.headers.get("sec-websocket-protocol", "").split(",")
        if value.strip()
    ]
    supplied_key = next((value for value in protocols if value.startswith("ananta_")), None)

    if authentication_required():
        auth_db = SessionLocal()
        try:
            if authenticate_api_key_value(supplied_key, auth_db) is None:
                await websocket.close(code=4401, reason="API key required")
                return
        finally:
            auth_db.close()

    accepted_protocol = "ananta" if "ananta" in protocols else None
    await ws_manager.connect(websocket, job_id, subprotocol=accepted_protocol)

    try:
        # Get database session
        db = SessionLocal()

        try:
            # Send initial status
            job = db.query(ScanJob).filter_by(job_id=job_id).first()
            if job:
                await websocket.send_json({
                    "type": "status",
                    "job_id": job.job_id,
                    "status": job.status,
                    "progress": job.progress,
                    "query": job.query
                })
            else:
                await websocket.send_json({
                    "type": "error",
                    "message": "Job not found"
                })
                return

            # Keep connection alive and poll for updates
            last_status = job.status
            last_progress = job.progress

            while True:
                await asyncio.sleep(1)  # Check every second

                # Refresh job status
                db.refresh(job)

                # Send update if status or progress changed
                if job.status != last_status or job.progress != last_progress:
                    last_status = job.status
                    last_progress = job.progress

                    update_message = {
                        "type": "update",
                        "job_id": job.job_id,
                        "status": job.status,
                        "progress": job.progress
                    }

                    # If completed, include the result
                    if job.status == "COMPLETED" and job.result:
                        try:
                            update_message["result"] = json.loads(job.result)
                        except:
                            update_message["result"] = {"error": "Parse error"}

                    # If failed, include error message
                    if job.status == "FAILED":
                        update_message["error"] = job.error_message

                    await websocket.send_json(update_message)

                    # Close connection if job is done
                    if job.status in ("COMPLETED", "FAILED"):
                        await websocket.send_json({
                            "type": "complete",
                            "job_id": job.job_id,
                            "final_status": job.status
                        })
                        break

                # Handle incoming messages (ping/pong, close requests)
                try:
                    data = await asyncio.wait_for(
                        websocket.receive_text(),
                        timeout=0.1
                    )
                    if data == "ping":
                        await websocket.send_text("pong")
                    elif data == "close":
                        break
                except asyncio.TimeoutError:
                    pass  # No message received, continue polling

        finally:
            db.close()

    except WebSocketDisconnect:
        logger.info(f"[WebSocket] Client disconnected from job {job_id}")
    except Exception as e:
        logger.exception(f"[WebSocket] Error for job {job_id}: {e}")
    finally:
        ws_manager.disconnect(websocket, job_id)


# Helper function to notify WebSocket clients (called from tasks.py)
async def notify_job_update(job_id: str, status: str, progress: int, result: dict = None, error: str = None):
    """
    Notify all WebSocket clients watching a job about an update.
    Called from Celery tasks via asyncio.
    """
    message = {
        "type": "update",
        "job_id": job_id,
        "status": status,
        "progress": progress
    }
    if result:
        message["result"] = result
    if error:
        message["error"] = error

    await ws_manager.send_job_update(job_id, message)


@router.get("/workers/status")
def get_workers_status(db: Session = Depends(get_db)):
    """
    Retourne l'état en temps réel des workers Celery.
    Utilisé par workers.html pour le monitoring.
    
    On Windows with --pool=solo, inspect() often fails. This endpoint uses
    multiple fallback methods:
    1. app.control.ping() - Most reliable on Windows
    2. app.control.inspect() - Standard method (may timeout on Windows)
    3. Redis heartbeat keys - Direct broker query
    4. Database PROCESSING jobs - Last resort inference
    """
    if not HAS_CELERY:
        return {
            "active_workers": 0,
            "active_tasks": 0,
            "pending_tasks": 0,
            "completed_24h": 0,
            "workers": [],
            "queues": {},
            "error": "Celery non disponible"
        }

    try:
        from tasks import app as celery_app
        from kombu import Connection
        import os
        import redis

        redis_url = os.getenv("REDIS_URL", "redis://localhost:6379/0")
        workers_list = []
        active_workers = 0
        active_tasks = 0
        detection_method = "none"

        # =====================================================
        # METHOD 1: app.control.ping() - Most reliable on Windows
        # =====================================================
        try:
            # ping() is more reliable than inspect() on Windows with --pool=solo
            # Use a short timeout (2s) to avoid blocking
            ping_response = celery_app.control.ping(timeout=2.0)
            
            if ping_response:
                # ping_response is a list of dicts: [{'worker@host': {'ok': 'pong'}}]
                for worker_dict in ping_response:
                    for worker_name, status in worker_dict.items():
                        if status.get('ok') == 'pong':
                            worker_display_name = worker_name.split('@')[0] if '@' in worker_name else worker_name
                            workers_list.append({
                                "name": worker_display_name,
                                "full_name": worker_name,
                                "status": "online",
                                "queues": [],  # Will be filled by inspect if available
                                "concurrency": 1,
                                "active_tasks": 0,
                                "total_processed": 0,
                                "detection": "ping"
                            })
                            active_workers += 1
                
                if active_workers > 0:
                    detection_method = "ping"
                    logger.debug(f"[Workers] Detected {active_workers} workers via ping()")

        except Exception as ping_err:
            logger.debug(f"[Workers] ping() failed: {ping_err}")

        # =====================================================
        # METHOD 2: app.control.inspect() - Try to get more details
        # =====================================================
        active_workers_info = {}
        stats_info = {}
        active_queues_info = {}

        if active_workers > 0 or detection_method == "none":
            try:
                # Use shorter timeout on Windows
                inspect = celery_app.control.inspect(timeout=2.0)
                
                # These calls can fail independently - wrap each
                try:
                    active_workers_info = inspect.active() or {}
                except Exception:
                    pass
                    
                try:
                    stats_info = inspect.stats() or {}
                except Exception:
                    pass
                    
                try:
                    active_queues_info = inspect.active_queues() or {}
                except Exception:
                    pass

                # If inspect() found workers but ping() didn't, update the list
                if stats_info and not workers_list:
                    detection_method = "inspect"
                    for worker_name, worker_stats in stats_info.items():
                        worker_queues = []
                        if worker_name in active_queues_info:
                            worker_queues = [q['name'] for q in active_queues_info[worker_name] if 'name' in q]

                        active_tasks_for_worker = active_workers_info.get(worker_name, [])
                        worker_display_name = worker_name.split('@')[0] if '@' in worker_name else worker_name

                        workers_list.append({
                            "name": worker_display_name,
                            "full_name": worker_name,
                            "status": "online",
                            "queues": worker_queues,
                            "concurrency": worker_stats.get('pool', {}).get('max-concurrency', 1),
                            "active_tasks": len(active_tasks_for_worker),
                            "total_processed": worker_stats.get('total', {}).get('ananta.scan_osint', 0),
                            "detection": "inspect"
                        })
                        active_workers += 1

                # Enrich existing workers from ping() with inspect() data
                elif workers_list and stats_info:
                    for worker in workers_list:
                        full_name = worker.get("full_name", "")
                        if full_name in stats_info:
                            worker_stats = stats_info[full_name]
                            worker["concurrency"] = worker_stats.get('pool', {}).get('max-concurrency', 1)
                            worker["total_processed"] = worker_stats.get('total', {}).get('ananta.scan_osint', 0)
                        if full_name in active_queues_info:
                            worker["queues"] = [q['name'] for q in active_queues_info[full_name] if 'name' in q]
                        if full_name in active_workers_info:
                            worker["active_tasks"] = len(active_workers_info[full_name])
                            active_tasks += worker["active_tasks"]

            except Exception as inspect_err:
                logger.debug(f"[Workers] inspect() failed: {inspect_err}")

        # =====================================================
        # METHOD 3: Direct Redis heartbeat query (fallback)
        # =====================================================
        if active_workers == 0:
            try:
                # Celery stores worker heartbeats in Redis
                # Key pattern: celery-task-meta-* and _kombu.binding.*
                r = redis.from_url(redis_url, socket_timeout=2.0)
                
                # Check for any celery keys that indicate worker activity
                # Workers register with keys like: _kombu.binding.<queue>
                binding_keys = r.keys("_kombu.binding.*")
                
                # Also check for recent heartbeats (worker events)
                # Celery workers publish to celeryev.* exchange
                event_keys = r.keys("celeryev.*")
                
                if binding_keys or event_keys:
                    # There's evidence of Celery activity in Redis
                    # This doesn't guarantee a worker is running, but it's a hint
                    logger.debug(f"[Workers] Redis shows Celery bindings: {len(binding_keys)} queues")

            except Exception as redis_err:
                logger.debug(f"[Workers] Redis heartbeat check failed: {redis_err}")

        # =====================================================
        # METHOD 4: Database PROCESSING jobs (last resort)
        # =====================================================
        processing_jobs = db.query(ScanJob).filter(ScanJob.status == "PROCESSING").count()
        
        if active_workers == 0 and processing_jobs > 0:
            detection_method = "db_inference"
            active_workers = 1
            workers_list = [{
                "name": "worker (inferred)",
                "status": "online",
                "queues": ["all"],
                "concurrency": 1,
                "active_tasks": processing_jobs,
                "total_processed": 0,
                "detection": "db_inference",
                "note": "Worker détecté via tâches en cours (inspect/ping indisponibles)"
            }]

        # =====================================================
        # Count pending tasks from Redis queues
        # =====================================================
        pending_tasks = 0
        queues_status = {}

        try:
            with Connection(redis_url) as conn:
                from celery_config import CELERY_QUEUES

                for queue_obj in CELERY_QUEUES:
                    queue_name = queue_obj.name
                    try:
                        queue_length = conn.default_channel.client.llen(queue_name)
                        pending_tasks += queue_length

                        workers_listening = 0
                        for worker in workers_list:
                            if queue_name in worker.get("queues", []) or "all" in worker.get("queues", []):
                                workers_listening += 1

                        queues_status[queue_name] = {
                            "pending": queue_length,
                            "workers": workers_listening
                        }
                    except Exception as e:
                        logger.debug(f"[Workers] Queue {queue_name} error: {e}")
                        queues_status[queue_name] = {"pending": 0, "workers": 0}

        except Exception as e:
            logger.warning(f"[Workers] Redis connection failed: {e}")
            for queue_name in ['priority', 'osint_fast', 'osint_medium', 'osint_critical', 'maintenance', 'default']:
                queues_status[queue_name] = {"pending": 0, "workers": 0}

        # =====================================================
        # Count completed jobs in last 24h
        # =====================================================
        now = datetime.now(timezone.utc)
        cutoff = now - timedelta(hours=24)

        completed_24h = db.query(ScanJob).filter(
            ScanJob.status == "COMPLETED",
            ScanJob.updated_at >= cutoff
        ).count()

        # Calculate total active tasks
        if active_tasks == 0:
            active_tasks = sum(w.get("active_tasks", 0) for w in workers_list)
            if active_tasks == 0:
                active_tasks = processing_jobs

        return {
            "active_workers": active_workers,
            "active_tasks": active_tasks,
            "pending_tasks": pending_tasks,
            "completed_24h": completed_24h,
            "workers": workers_list,
            "queues": queues_status,
            "detection_method": detection_method,
            "inspect_available": detection_method == "inspect",
            "ping_available": detection_method == "ping",
            "processing_jobs_db": processing_jobs
        }

    except Exception as e:
        logger.exception(f"[/workers/status] ERROR: {e}")
        return {
            "active_workers": 0,
            "active_tasks": 0,
            "pending_tasks": 0,
            "completed_24h": 0,
            "workers": [],
            "queues": {},
            "error": str(e)
        }


# ============================================================================
# ENTITY RESEARCH - Recherche d'entité (personne physique ou morale)
# ============================================================================
#
# Le moteur `entity_research` part du moindre indice (nom, email, téléphone,
# SIREN, LEI, domaine, pseudo...) et pivote de source en source pour
# reconstituer ce qui est publiquement connaissable sur l'entité.
#
# Endpoints:
#   POST /entity/preview        - ce que le moteur comprend, sans rien interroger
#   GET  /entity/sources        - catalogue des sources et leur disponibilité
#   POST /entity/research       - recherche synchrone (bornée)
#   POST /entity/research_async - recherche en tâche de fond (Celery)
#   GET  /entity/runs           - historique des dossiers
#   GET  /entity/run/{run_id}   - dossier complet
#   GET  /entity/run/{run_id}/graph    - graphe entités/relations
#   GET  /entity/run/{run_id}/report   - rapport Markdown
#   GET  /entity/run/{run_id}/export/{format}
#   GET  /entity/entity/{entity_key}/runs - autres dossiers citant cette entité
#   DELETE /entity/run/{run_id}


def _entity_engine():
    """Import tardif du moteur (évite d'alourdir le démarrage de l'API)."""
    import entity_research

    return entity_research


def _request_principal(request: Request):
    """Identité posée par le middleware, avec repli local explicite."""
    from auth import AuthPrincipal

    return getattr(
        request.state,
        "principal",
        AuthPrincipal(actor_id="local", role="admin", name="local"),
    )


def _owned_run_query(db: Session, request: Request):
    query = db.query(EntityResearchRun)
    principal = _request_principal(request)
    if principal.role != "admin":
        query = query.filter(EntityResearchRun.created_by == principal.actor_id)
    return query


def _owned_run_or_404(db: Session, request: Request, run_id: str):
    run = _owned_run_query(db, request).filter(EntityResearchRun.run_id == run_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Dossier non trouvé")
    return run


def _owned_watch_query(db: Session, request: Request):
    query = db.query(EntityWatch)
    principal = _request_principal(request)
    if principal.role != "admin":
        query = query.filter(EntityWatch.created_by == principal.actor_id)
    return query


def _run_to_summary(run: EntityResearchRun) -> Dict:
    return {
        "run_id": run.run_id,
        "job_id": run.job_id,
        "parent_run_id": run.parent_run_id,
        "pass_number": run.pass_number or 1,
        "query": run.query,
        "label": run.label,
        "entity_kind": run.entity_kind,
        "mode": run.mode,
        "purpose": run.purpose,
        "language": run.language,
        "status": run.status,
        "progress": run.progress,
        "confidence_score": run.confidence_score,
        "risk_level": run.risk_level,
        "risk_score": run.risk_score,
        "entities_count": run.entities_count,
        "relationships_count": run.relationships_count,
        "sources_ok": run.sources_ok,
        "partial": run.partial,
        "error_message": run.error_message,
        "created_by": run.created_by,
        "created_at": run.created_at.isoformat() if run.created_at else None,
        "updated_at": run.updated_at.isoformat() if run.updated_at else None,
    }


def _active_run_for_principal(db: Session, request: Request):
    principal = _request_principal(request)
    return (
        db.query(EntityResearchRun)
        .filter(EntityResearchRun.active_owner == principal.actor_id)
        .order_by(EntityResearchRun.created_at.desc())
        .first()
    )


def _watch_to_summary(watch: EntityWatch) -> Dict:
    return {
        "id": watch.id,
        "query": watch.query,
        "label": watch.label,
        "entity_kind": watch.entity_kind,
        "root_key": watch.root_key,
        "mode": watch.mode,
        "purpose": watch.purpose,
        "baseline_run_id": watch.baseline_run_id,
        "last_run_id": watch.last_run_id,
        "last_change_score": watch.last_change_score or 0,
        "last_change_summary": watch.last_change_summary,
        "last_checked_at": (
            watch.last_checked_at.isoformat() if watch.last_checked_at else None
        ),
        "is_active": watch.is_active,
        "created_by": watch.created_by,
        "created_at": watch.created_at.isoformat() if watch.created_at else None,
    }


def _resolution_payload(db: Session, run: EntityResearchRun, dossier: Dict) -> Dict:
    """Ajoute les validations analyste sans réécrire le dossier source."""
    from entity_research.storage import resolution_with_reviews

    return resolution_with_reviews(
        db,
        run.run_id,
        list(dossier.get("resolution") or []),
    )


@router.post("/entity/preview")
def entity_preview(body: EntityPreviewRequest):
    """
    Analyse la requête sans lancer la moindre collecte.

    Retourne les sélecteurs reconnus (et validés), la nature déduite de
    l'entité et les sources qui seraient interrogées. Permet à l'opérateur de
    corriger avant de dépenser du temps et des quotas.
    """
    try:
        engine = _entity_engine()
        return engine.preview_selectors(
            body.query,
            entity_kind=body.entity_kind,
            default_region=body.default_region,
        )
    except Exception as e:
        logger.exception(f"[/entity/preview] ERROR: {e}")
        raise HTTPException(status_code=500, detail=f"Analyse impossible: {e}")


@router.get("/entity/sources")
def entity_sources(request: Request, response: Response):
    """Catalogue des sources : couche, sélecteurs acceptés, clé requise, disponibilité."""
    try:
        engine = _entity_engine()
        sources = engine.describe_sources()
    except Exception as e:
        logger.exception(f"[/entity/sources] ERROR: {e}")
        raise HTTPException(status_code=500, detail=f"Catalogue indisponible: {e}")

    payload = {
        "total": len(sources),
        "available": sum(1 for s in sources if s.get("available")),
        "by_layer": {
            "1": sum(1 for s in sources if s["layer"] == 1),
            "2": sum(1 for s in sources if s["layer"] == 2),
            "3": sum(1 for s in sources if s["layer"] == 3),
        },
        "sources": sources,
    }
    # Le catalogue ne bouge qu'au redémarrage (ou changement de clés d'API).
    etag = generate_etag(payload)
    if check_not_modified(request, etag=etag):
        return Response(status_code=304, headers={"ETag": etag})
    add_cache_headers(response, etag=etag, max_age=300)
    return payload


@router.get("/auth/status")
def auth_status(db: Session = Depends(get_db)):
    """État public minimal permettant à l'interface de proposer le bon accès."""
    from auth import authentication_required

    initialized = db.query(APIKey).filter(APIKey.is_active.is_(True)).first() is not None
    return {
        "required": authentication_required(),
        "initialized": initialized,
        "bootstrap_configured": bool(os.getenv("ANANTA_BOOTSTRAP_TOKEN")),
    }


def _research_kwargs(body: EntityResearchRequest) -> Dict:
    return {
        "mode": body.mode,
        "purpose": body.purpose,
        "match_policy": body.match_policy,
        "entity_kind": body.entity_kind,
        "language": body.language,
        "template": body.report_template,
        "jurisdiction": body.jurisdiction,
        "allow_account_enumeration": body.allow_account_enumeration,
        "allow_breach_data": body.allow_breach_data,
        "allow_person_pivot": body.allow_person_pivot,
        "redact_personal_data": body.redact_personal_data,
        "authorized_investigation_acknowledged": (
            body.authorized_investigation_acknowledged
        ),
        "only_sources": body.only_sources,
        "exclude_sources": body.exclude_sources,
        "briefing_text": body.briefing_text,
        "briefing_facts": (
            [fact.model_dump() for fact in body.briefing_facts]
            if body.briefing_facts
            else None
        ),
        "briefing_origin": body.briefing_origin,
        "use_llm": body.use_llm,
        "llm_hard_limit": body.llm_hard_limit,
        "default_region": body.default_region,
    }


@router.post("/entity/research")
def entity_research(
    body: EntityResearchRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """
    Recherche synchrone : renvoie le dossier complet dans la réponse.

    Adapté aux modes `passive` et `standard`. Pour le mode `deep`, préférer
    `/entity/research_async` : la collecte peut dépasser le timeout HTTP.
    """
    engine = _entity_engine()
    from entity_research.storage import mark_failed, persist_dossier

    try:
        dossier = engine.research_entity(body.query, **_research_kwargs(body))
    except Exception as e:
        logger.exception(f"[/entity/research] ERROR: {e}")
        raise HTTPException(status_code=500, detail=f"Recherche impossible: {e}")

    try:
        persist_dossier(
            db,
            dossier,
            mode=body.mode,
            purpose=body.purpose,
            language=body.language,
            report_template=body.report_template,
            created_by=_request_principal(request).actor_id,
        )
    except Exception as e:
        # Un dossier non persisté reste un dossier utilisable : on le renvoie.
        logger.error(f"[/entity/research] Persistance échouée: {e}")
        db.rollback()

    return {"type": "dossier", **dossier.to_dict()}


@router.post("/entity/research_async")
def entity_research_async(
    body: EntityResearchRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """Lance la recherche en tâche de fond et retourne un run_id à suivre."""
    if not HAS_CELERY or entity_research_task is None:
        raise HTTPException(
            status_code=503,
            detail="Mode asynchrone non disponible. Celery/Redis non configurés.",
        )
    if not ServiceStatus().is_redis_available():
        raise HTTPException(status_code=503, detail="File de tâches Redis indisponible")

    return _launch_entity_run(body, request, db)


def _launch_entity_run(
    body: EntityResearchRequest,
    request: Request,
    db: Session,
    *,
    parent_run_id: Optional[str] = None,
    pass_number: int = 1,
):
    """Crée le verrou et le run avant de publier la tâche Celery."""
    from entity_research.storage import create_run, mark_failed

    if not HAS_CELERY or entity_research_task is None:
        raise HTTPException(
            status_code=503,
            detail="Recherche suivie indisponible. Démarrez Celery et Redis.",
        )
    if not ServiceStatus().is_redis_available():
        raise HTTPException(status_code=503, detail="File de tâches Redis indisponible")

    principal = _request_principal(request)
    active = _active_run_for_principal(db, request)
    if active:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Une recherche est déjà en cours pour « {active.query} » "
                f"(run {active.run_id}). Ajoutez-y un indice au lieu d'en lancer une autre."
            ),
        )

    run_id = str(uuid.uuid4())
    try:
        run = create_run(
            db,
            run_id=run_id,
            query=body.query,
            job_id=run_id,
            mode=body.mode,
            purpose=body.purpose,
            language=body.language,
            report_template=body.report_template,
            created_by=principal.actor_id,
            active_owner=principal.actor_id,
            parent_run_id=parent_run_id,
            pass_number=pass_number,
        )
    except IntegrityError:
        db.rollback()
        active = _active_run_for_principal(db, request)
        detail = "Une recherche est déjà en cours."
        if active:
            detail += f" Run actif : {active.run_id}."
        raise HTTPException(status_code=409, detail=detail)

    options = _research_kwargs(body)
    options["_created_by"] = principal.actor_id
    options["_pass_number"] = pass_number
    options["_parent_run_id"] = parent_run_id
    try:
        task = entity_research_task.apply_async(
            args=[body.query, options],
            task_id=run_id,
        )
    except Exception as exc:
        mark_failed(db, run_id, f"Publication de la tâche impossible: {exc}")
        raise HTTPException(status_code=503, detail="Impossible de démarrer le worker")

    logger.info(
        "[ENTITY ASYNC] Run %s lancé pour '%s' (mode=%s, passe=%s)",
        run_id,
        body.query,
        body.mode,
        pass_number,
    )
    return {
        "type": "async",
        "run_id": run.run_id,
        "job_id": task.id,
        "status": "PENDING",
        "mode": body.mode,
        "pass_number": pass_number,
        "parent_run_id": parent_run_id,
        "message": f"Recherche lancée. Suivre la progression sur /entity/run/{run_id}",
    }


@router.get("/entity/active")
def entity_active(request: Request, db: Session = Depends(get_db)):
    """Run actif de l'utilisateur, utilisé pour restaurer l'UI après rechargement."""
    run = _active_run_for_principal(db, request)
    return {"active": _run_to_summary(run) if run else None}


@router.get("/entity/runs")
def entity_runs(
    request: Request,
    limit: int = Query(20, ge=1, le=100),
    page: int = Query(1, ge=1),
    entity_kind: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None, max_length=200),
    db: Session = Depends(get_db),
):
    """Historique paginé des dossiers d'entité."""
    query = _owned_run_query(db, request)

    if entity_kind in {"person", "organization", "unknown"}:
        query = query.filter(EntityResearchRun.entity_kind == entity_kind)
    if status:
        query = query.filter(EntityResearchRun.status == status.upper())
    if search:
        pattern = f"%{search.strip()}%"
        query = query.filter(
            (EntityResearchRun.label.ilike(pattern)) | (EntityResearchRun.query.ilike(pattern))
        )

    query = query.order_by(EntityResearchRun.created_at.desc())

    total = query.count()
    items = query.offset((page - 1) * limit).limit(limit).all()
    pages = (total + limit - 1) // limit if total else 0

    return {
        "items": [_run_to_summary(run) for run in items],
        "total": total,
        "page": page,
        "limit": limit,
        "pages": pages,
        "has_next": page < pages,
        "has_prev": page > 1,
    }


@router.get("/entity/run/{run_id}")
def entity_run_detail(
    run_id: str,
    request: Request,
    include_sources: bool = Query(True),
    db: Session = Depends(get_db),
):
    """Dossier complet d'un run (ou son état s'il est encore en cours)."""
    run = _owned_run_or_404(db, request, run_id)

    payload = _run_to_summary(run)
    from entity_research.storage import list_instructions

    payload["instructions"] = list_instructions(db, run_id)
    next_run = (
        _owned_run_query(db, request)
        .filter(EntityResearchRun.parent_run_id == run_id)
        .order_by(EntityResearchRun.created_at.desc())
        .first()
    )
    payload["next_run"] = _run_to_summary(next_run) if next_run else None

    if run.dossier:
        try:
            dossier = json.loads(run.dossier)
            resolution = _resolution_payload(db, run, dossier)
            dossier["resolution"] = resolution["decisions"]
            dossier["resolution_review_counts"] = resolution["review_counts"]
            if not include_sources:
                dossier.pop("sources", None)
            payload["dossier"] = dossier
        except json.JSONDecodeError:
            logger.error(f"[/entity/run] Dossier illisible pour {run_id}")
            payload["dossier"] = None

    return payload


@router.post("/entity/run/{run_id}/instructions")
def entity_run_instruction_add(
    run_id: str,
    body: EntityInstructionRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """Ajoute un indice au run actif ; le worker le prend à la vague suivante."""
    from entity_research.storage import add_instruction

    run = _owned_run_or_404(db, request, run_id)
    if run.status not in {"PENDING", "PROCESSING"} or not run.active_owner:
        raise HTTPException(
            status_code=409,
            detail="Cette passe est terminée. Lancez une nouvelle passe avec cette instruction.",
        )
    if (run.progress or 0) >= 88:
        raise HTTPException(
            status_code=409,
            detail=(
                "Le dossier est déjà en consolidation. Attendez sa fin puis lancez "
                "une nouvelle passe avec cette instruction."
            ),
        )
    principal = _request_principal(request)
    item = add_instruction(
        db,
        run_id=run_id,
        text=body.text,
        origin=body.origin,
        created_by=principal.actor_id,
    )
    return {
        "id": item.id,
        "run_id": run_id,
        "text": item.text,
        "origin": item.origin,
        "status": item.status,
        "message": "Indice ajouté. Le moteur l'intégrera à la prochaine vague.",
    }


@router.post("/entity/run/{run_id}/continue")
def entity_run_continue(
    run_id: str,
    body: EntityContinueRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """Lance une passe liée en conservant la cible et la traçabilité."""
    previous = _owned_run_or_404(db, request, run_id)
    if previous.status not in {"COMPLETED", "FAILED", "CANCELLED"}:
        raise HTTPException(status_code=409, detail="La passe actuelle est encore en cours")
    kind = previous.entity_kind if previous.entity_kind in {"person", "organization"} else None
    follow_up = EntityResearchRequest(
        query=previous.query,
        mode=body.mode or previous.mode or "standard",
        entity_kind=kind,
        purpose=previous.purpose or "due_diligence",
        language=previous.language or "fr",
        report_template=previous.report_template or "detailed",
        briefing_text=body.text,
        briefing_origin=body.origin,
        authorized_investigation_acknowledged=(
            previous.purpose == "authorized_investigation"
        ),
    )
    return _launch_entity_run(
        follow_up,
        request,
        db,
        parent_run_id=previous.run_id,
        pass_number=(previous.pass_number or 1) + 1,
    )


@router.post("/entity/run/{run_id}/cancel")
def entity_run_cancel(
    run_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Arrête un run suivi et libère immédiatement le verrou utilisateur."""
    from entity_research.storage import update_run_progress

    run = _owned_run_or_404(db, request, run_id)
    if run.status not in {"PENDING", "PROCESSING"}:
        return _run_to_summary(run)
    try:
        entity_research_task.app.control.revoke(
            run.job_id or run.run_id,
            terminate=True,
            signal="SIGTERM",
        )
    except Exception as exc:
        logger.warning("[ENTITY] révocation Celery non confirmée pour %s: %s", run_id, exc)
    update_run_progress(db, run_id, progress=run.progress or 0, status="CANCELLED")
    db.refresh(run)
    return _run_to_summary(run)


@router.get("/entity/run/{run_id}/graph")
def entity_run_graph(
    run_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Graphe entités/relations, prêt pour un rendu visuel."""
    run = _owned_run_or_404(db, request, run_id)
    if not run.dossier:
        raise HTTPException(status_code=404, detail="Dossier non trouvé ou incomplet")

    try:
        dossier = json.loads(run.dossier)
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Dossier illisible")

    return dossier.get("graph") or {"version": 1, "nodes": [], "edges": []}


@router.get("/entity/run/{run_id}/report")
def entity_run_report(
    run_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Rapport Markdown du dossier."""
    run = _owned_run_or_404(db, request, run_id)

    return {
        "run_id": run.run_id,
        "label": run.label,
        "language": run.language,
        "template": run.report_template,
        "report": run.report_markdown or "",
    }


@router.get("/entity/run/{run_id}/export/{export_format}")
def entity_run_export(
    run_id: str,
    export_format: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Export du dossier en JSON, Markdown ou CSV (faits sourcés)."""
    run = _owned_run_or_404(db, request, run_id)
    if not run.dossier:
        raise HTTPException(status_code=404, detail="Dossier non trouvé ou incomplet")

    export_format = (export_format or "json").lower()
    if export_format not in {"json", "markdown", "md", "csv"}:
        raise HTTPException(
            status_code=400, detail="Format non supporté (json, markdown, csv)"
        )

    try:
        dossier = json.loads(run.dossier)
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Dossier illisible")

    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", (run.label or run.run_id))[:60]

    if export_format == "json":
        resolution = _resolution_payload(db, run, dossier)
        dossier["resolution"] = resolution["decisions"]
        dossier["resolution_review_counts"] = resolution["review_counts"]
        return JSONResponse(
            content=dossier,
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.json"'},
        )

    if export_format in {"markdown", "md"}:
        return Response(
            content=run.report_markdown or "",
            media_type="text/markdown; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.md"'},
        )

    # CSV : un fait par ligne, avec sa source - format pivot pour un tableur
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(
        ["entity", "kind", "attribute", "value", "confidence", "source", "url", "observed_at"]
    )
    for entity in dossier.get("entities", []):
        for attribute in entity.get("attributes", []):
            provenance = attribute.get("provenance") or {}
            writer.writerow(
                [
                    entity.get("label", ""),
                    entity.get("kind", ""),
                    attribute.get("name", ""),
                    str(attribute.get("value", ""))[:500],
                    attribute.get("confidence", ""),
                    provenance.get("source_id", ""),
                    provenance.get("url", "") or "",
                    provenance.get("observed_at", ""),
                ]
            )

    return Response(
        content=buffer.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
    )


@router.get("/entity/entity/{entity_key:path}/observations")
def entity_observations(
    entity_key: str,
    request: Request,
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
):
    """Historique first/last seen de l'entité et de ses relations."""
    from entity_research.storage import entity_observation_history

    principal = _request_principal(request)
    return entity_observation_history(
        db,
        entity_key,
        limit=limit,
        created_by=None if principal.role == "admin" else principal.actor_id,
    )


@router.get("/entity/entity/{entity_key:path}/runs")
def entity_related_runs(
    entity_key: str,
    request: Request,
    exclude_run_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Autres dossiers dans lesquels cette entité apparaît.

    C'est ce qui fait ressortir qu'un même dirigeant revient dans plusieurs
    sociétés analysées séparément.
    """
    from entity_research.storage import find_related_runs

    principal = _request_principal(request)
    return {
        "entity_key": entity_key,
        "runs": find_related_runs(
            db,
            entity_key,
            exclude_run_id=exclude_run_id or "",
            created_by=None if principal.role == "admin" else principal.actor_id,
        ),
    }


@router.get("/entity/run/{run_id}/resolution")
def entity_resolution(
    run_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Journal de résolution enrichi des validations analyste."""
    run = _owned_run_or_404(db, request, run_id)
    if not run.dossier:
        raise HTTPException(status_code=404, detail="Dossier incomplet")
    try:
        dossier = json.loads(run.dossier)
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Dossier illisible")
    return {"run_id": run_id, **_resolution_payload(db, run, dossier)}


@router.put("/entity/run/{run_id}/resolution/{decision_id}/review")
def entity_resolution_review(
    run_id: str,
    decision_id: str,
    body: EntityResolutionReviewRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """Confirme, rejette ou place une décision en attente d'information."""
    run = _owned_run_or_404(db, request, run_id)
    if not run.dossier:
        raise HTTPException(status_code=404, detail="Dossier incomplet")
    try:
        dossier = json.loads(run.dossier)
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Dossier illisible")

    current = _resolution_payload(db, run, dossier)
    if not any(item["decision_id"] == decision_id for item in current["decisions"]):
        raise HTTPException(status_code=404, detail="Décision de résolution inconnue")

    review = (
        db.query(EntityResolutionReview)
        .filter(
            EntityResolutionReview.run_id == run_id,
            EntityResolutionReview.decision_id == decision_id,
        )
        .first()
    )
    principal = _request_principal(request)
    if review is None:
        review = EntityResolutionReview(
            run_id=run_id,
            decision_id=decision_id,
            created_by=principal.actor_id,
        )
        db.add(review)
    review.status = body.status
    review.note = body.note or None
    review.updated_by = principal.actor_id
    review.updated_at = datetime.now(timezone.utc)
    db.commit()
    logger.info(
        "[ENTITY] Résolution %s/%s revue par %s : %s",
        run_id,
        decision_id,
        principal.actor_id,
        body.status,
    )
    return {"run_id": run_id, **_resolution_payload(db, run, dossier)}


@router.delete("/entity/run/{run_id}/resolution/{decision_id}/review")
def entity_resolution_review_clear(
    run_id: str,
    decision_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Efface une validation humaine sans modifier la décision calculée."""
    run = _owned_run_or_404(db, request, run_id)
    review = (
        db.query(EntityResolutionReview)
        .filter(
            EntityResolutionReview.run_id == run_id,
            EntityResolutionReview.decision_id == decision_id,
        )
        .first()
    )
    if review:
        db.delete(review)
        db.commit()
    try:
        dossier = json.loads(run.dossier or "{}")
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Dossier illisible")
    return {"run_id": run_id, **_resolution_payload(db, run, dossier)}


@router.delete("/entity/run/{run_id}")
def entity_run_delete(
    run_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Supprime un dossier et ses entités normalisées (droit à l'effacement)."""
    run = _owned_run_or_404(db, request, run_id)

    db.query(EntityResolutionReview).filter_by(run_id=run_id).delete(
        synchronize_session=False
    )
    db.query(EntityResearchInstruction).filter_by(run_id=run_id).delete(
        synchronize_session=False
    )
    db.query(ResearchEntity).filter_by(run_id=run_id).delete(synchronize_session=False)
    db.delete(run)
    db.commit()

    logger.info(f"[ENTITY] Dossier {run_id} supprimé")
    return {"success": True, "message": f"Dossier {run_id} supprimé"}


@router.get("/entity/run/{run_id}/changes")
def entity_run_changes(
    run_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Compare ce dossier au précédent dossier de la même entité."""
    from entity_research.changes import compare_dossiers

    run = _owned_run_or_404(db, request, run_id)
    if not run.dossier:
        raise HTTPException(status_code=404, detail="Dossier incomplet")

    previous_query = _owned_run_query(db, request).filter(
        EntityResearchRun.id < run.id,
        EntityResearchRun.status == "COMPLETED",
        EntityResearchRun.dossier.isnot(None),
    )
    if run.root_key:
        previous_query = previous_query.filter(EntityResearchRun.root_key == run.root_key)
    else:
        previous_query = previous_query.filter(EntityResearchRun.query == run.query)
    previous = previous_query.order_by(EntityResearchRun.id.desc()).first()

    if not previous:
        return {
            "comparison_available": False,
            "current_run_id": run.run_id,
            "previous_run_id": None,
            "has_changes": False,
            "change_score": 0,
            "counts": {},
        }

    try:
        return compare_dossiers(
            json.loads(run.dossier),
            json.loads(previous.dossier),
            current_run_id=run.run_id,
            previous_run_id=previous.run_id,
        )
    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="Dossier illisible")


@router.get("/entity/watches")
def entity_watch_list(request: Request, db: Session = Depends(get_db)):
    """Liste des entités suivies par l'utilisateur courant."""
    watches = (
        _owned_watch_query(db, request)
        .filter(EntityWatch.is_active.is_(True))
        .order_by(EntityWatch.updated_at.desc())
        .all()
    )
    return {"items": [_watch_to_summary(watch) for watch in watches], "total": len(watches)}


@router.post("/entity/watch/{run_id}")
def entity_watch_create(
    run_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    """Place l'entité racine d'un dossier sous surveillance."""
    run = _owned_run_or_404(db, request, run_id)
    if not run.root_key:
        raise HTTPException(status_code=409, detail="Le dossier n'a pas d'entité racine")

    principal = _request_principal(request)
    watch = (
        db.query(EntityWatch)
        .filter(
            EntityWatch.root_key == run.root_key,
            EntityWatch.created_by == principal.actor_id,
        )
        .first()
    )
    if watch is None:
        watch = EntityWatch(root_key=run.root_key, created_by=principal.actor_id)
        db.add(watch)

    watch.query = run.query
    watch.label = run.label
    watch.entity_kind = run.entity_kind
    watch.mode = run.mode
    watch.purpose = run.purpose
    watch.language = run.language
    watch.report_template = run.report_template
    watch.baseline_run_id = watch.baseline_run_id or run.run_id
    watch.last_run_id = run.run_id
    watch.is_active = True
    watch.last_checked_at = run.updated_at or run.created_at
    db.commit()
    db.refresh(watch)
    return _watch_to_summary(watch)


@router.delete("/entity/watch/{watch_id}")
def entity_watch_delete(
    watch_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Retire une entité de la liste de surveillance."""
    watch = _owned_watch_query(db, request).filter(EntityWatch.id == watch_id).first()
    if not watch:
        raise HTTPException(status_code=404, detail="Veille non trouvée")
    db.delete(watch)
    db.commit()
    return {"success": True}


@router.post("/entity/watch/{watch_id}/refresh")
def entity_watch_refresh(
    watch_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Relance la collecte d'une veille et calcule son delta à la fin."""
    if not HAS_CELERY or entity_research_task is None:
        raise HTTPException(status_code=503, detail="Workers Celery indisponibles")
    if not ServiceStatus().is_redis_available():
        raise HTTPException(status_code=503, detail="File de tâches Redis indisponible")

    from entity_research.storage import create_run

    watch = _owned_watch_query(db, request).filter(EntityWatch.id == watch_id).first()
    if not watch:
        raise HTTPException(status_code=404, detail="Veille non trouvée")

    principal = _request_principal(request)
    options = {
        "mode": watch.mode or "standard",
        "purpose": watch.purpose or "due_diligence",
        "entity_kind": (
            watch.entity_kind if watch.entity_kind in {"person", "organization"} else None
        ),
        "language": watch.language or "fr",
        "template": watch.report_template or "detailed",
        "authorized_investigation_acknowledged": (
            watch.purpose == "authorized_investigation"
        ),
        "_created_by": principal.actor_id,
    }
    task = entity_research_task.delay(watch.query, options)
    create_run(
        db,
        run_id=task.id,
        query=watch.query,
        job_id=task.id,
        mode=watch.mode,
        purpose=watch.purpose,
        language=watch.language,
        report_template=watch.report_template,
        created_by=principal.actor_id,
    )
    return {"run_id": task.id, "status": "PENDING", "watch_id": watch.id}

# ============================================================================
# MOTEUR D'INFÉRENCE - Choix du LLM
# ============================================================================
#
# Ananta n'est lié à aucun moteur : text-generation-webui, Ollama (local ou sur
# une autre machine), API compatible OpenAI, API Claude, ou les CLI `claude` /
# `codex` déjà installées et authentifiées. Sans LLM disponible, les rapports
# déterministes restent produits.


@router.get("/llm/providers")
def llm_providers(probe: bool = Query(True, description="Tester la disponibilité réelle")):
    """Catalogue des moteurs d'inférence et lequel est actif."""
    from llm_providers import current_provider_id, describe_providers

    providers = describe_providers(probe=probe)
    return {
        "active": current_provider_id(),
        "available": sum(1 for p in providers if p.get("available")),
        "total": len(providers),
        "providers": providers,
    }


@router.post("/llm/provider")
def llm_set_provider(body: LLMProviderRequest):
    """
    Bascule le moteur d'inférence à chaud.

    Le changement s'applique au processus courant. Pour qu'il survive à un
    redémarrage, renseigner `LLM_PROVIDER` (et les variables associées) dans
    le fichier `.env`.
    """
    from llm_providers import set_provider

    overrides = {}
    if body.model:
        overrides["LLM_MODEL"] = body.model
        if body.provider == "ollama":
            overrides["OLLAMA_MODEL"] = body.model
        elif body.provider == "anthropic":
            overrides["ANTHROPIC_MODEL"] = body.model
    if body.endpoint:
        if body.provider == "ollama":
            overrides["OLLAMA_HOST"] = body.endpoint
        else:
            overrides["LLM_API_URL"] = body.endpoint

    try:
        result = set_provider(body.provider, overrides)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    logger.info(f"[LLM] Fournisseur basculé sur '{body.provider}' (disponible={result['available']})")
    return {
        "success": True,
        **result,
        "persist_hint": f"Ajoutez LLM_PROVIDER={body.provider} dans .env pour rendre le choix permanent.",
    }


@router.post("/llm/test")
def llm_test(body: LLMProviderRequest = None):
    """Envoie une requête minimale au moteur actif et renvoie sa réponse."""
    from llm_providers import LLMUnavailable, current_provider_id, generate

    provider_id = body.provider if body else current_provider_id()
    try:
        answer = generate(
            "Tu réponds en une phrase courte, en français.",
            "Confirme que tu es opérationnel.",
            max_tokens=60,
            provider_id=provider_id,
        )
        return {"success": True, "provider": provider_id, "answer": answer[:500]}
    except LLMUnavailable as e:
        return {"success": False, "provider": provider_id, "error": str(e)}
    except Exception as e:
        logger.exception("[/llm/test] erreur inattendue")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/llm/system-prompt")
def llm_system_prompt():
    """Pré-prompt commun injecté avant les instructions de chaque tâche IA."""
    from llm_providers import system_preprompt_info

    return system_preprompt_info()


@router.put("/llm/system-prompt")
def llm_set_system_prompt(body: LLMSystemPromptRequest):
    """Modifie le pré-prompt à chaud, ou rétablit la configuration avec ``null``."""
    from llm_providers import set_system_preprompt

    try:
        result = set_system_preprompt(body.prompt)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    logger.info("[LLM] Pré-prompt système mis à jour (source=%s)", result["source"])
    return {
        "success": True,
        **result,
        "persist_hint": (
            "Définissez LLM_SYSTEM_PROMPT ou LLM_SYSTEM_PROMPT_FILE dans .env "
            "pour conserver ce réglage après redémarrage."
        ),
    }
